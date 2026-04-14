from flask import Flask, request, jsonify, render_template, send_file, send_from_directory, session, redirect, url_for, make_response, g
import json, os, random, string, zipfile, io, tempfile, csv, re
import logging, time, uuid
from html import escape
from datetime import datetime, timedelta, timezone
from collections import deque
from functools import wraps
from logging.handlers import RotatingFileHandler
from urllib import request as urllib_request, parse as urllib_parse, error as urllib_error
from PIL import Image, ImageOps, ImageEnhance, ImageDraw, UnidentifiedImageError
from werkzeug.utils import secure_filename

app = Flask(__name__)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUNTIME_DIR = tempfile.gettempdir() if os.environ.get("VERCEL") else BASE_DIR
UPLOAD_DIR = os.path.join(RUNTIME_DIR, "static", "uploads")
ASSET_DIR = os.path.join(RUNTIME_DIR, "static", "assets")
DATA_FILE = os.path.join(RUNTIME_DIR, "records.json")
CERTIFICATE_DATA_FILE = os.path.join(RUNTIME_DIR, "certificates.json")
SETTINGS_FILE = os.path.join(RUNTIME_DIR, "settings.json")
ADMIN_PREFS_FILE = os.path.join(RUNTIME_DIR, "admin_prefs.json")
STORE_DIR = os.path.join(RUNTIME_DIR, "data_store")
RECORDS_DIR = os.path.join(STORE_DIR, "records")
CERTIFICATES_DIR = os.path.join(STORE_DIR, "certificates")
SETTINGS_DIR = os.path.join(STORE_DIR, "settings")
LOGS_DIR = os.path.join(RUNTIME_DIR, "logs")
APP_LOG_FILE = os.path.join(LOGS_DIR, "app.log")
AUDIT_LOG_FILE = os.path.join(LOGS_DIR, "audit.log")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(ASSET_DIR, exist_ok=True)
os.makedirs(RECORDS_DIR, exist_ok=True)
os.makedirs(CERTIFICATES_DIR, exist_ok=True)
os.makedirs(SETTINGS_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)
app.config["SIGNATURE_UPLOAD_PASSWORD"] = os.environ.get("SIGNATURE_UPLOAD_PASSWORD", "admin123")
app.config["ADMIN_PANEL_PASSWORD"] = os.environ.get("ADMIN_PANEL_PASSWORD", "admin123")
app.config["SUPABASE_URL"] = os.environ.get("SUPABASE_URL", "").strip()
app.config["SUPABASE_SERVICE_ROLE_KEY"] = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
app.config["SUPABASE_PHOTOS_BUCKET"] = os.environ.get("SUPABASE_PHOTOS_BUCKET", "id-card-photos").strip() or "id-card-photos"
app.config["APP_BUILD_TAG"] = os.environ.get("APP_BUILD_TAG", "").strip() or datetime.utcnow().strftime("%Y%m%d-%H%M%S")
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.config["CRON_BACKUP_TOKEN"] = os.environ.get("CRON_BACKUP_TOKEN", "").strip()
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "medical-id-card-secret")
HEX_COLOR_PATTERN = re.compile(r"^#[0-9a-fA-F]{6}$")


def configure_project_logging():
    formatter = logging.Formatter("%(asctime)s %(levelname)s %(name)s %(message)s")
    has_app_file_handler = any(
        isinstance(handler, RotatingFileHandler) and getattr(handler, "baseFilename", "") == APP_LOG_FILE
        for handler in app.logger.handlers
    )
    if not has_app_file_handler:
        app_file_handler = RotatingFileHandler(APP_LOG_FILE, maxBytes=2 * 1024 * 1024, backupCount=5, encoding="utf-8")
        app_file_handler.setLevel(logging.INFO)
        app_file_handler.setFormatter(formatter)
        app.logger.addHandler(app_file_handler)
    app.logger.setLevel(logging.INFO)

    audit_logger = logging.getLogger("medical_id.audit")
    audit_logger.setLevel(logging.INFO)
    audit_logger.propagate = False
    has_audit_handler = any(
        isinstance(handler, RotatingFileHandler) and getattr(handler, "baseFilename", "") == AUDIT_LOG_FILE
        for handler in audit_logger.handlers
    )
    if not has_audit_handler:
        audit_handler = RotatingFileHandler(AUDIT_LOG_FILE, maxBytes=4 * 1024 * 1024, backupCount=10, encoding="utf-8")
        audit_handler.setLevel(logging.INFO)
        audit_handler.setFormatter(logging.Formatter("%(message)s"))
        audit_logger.addHandler(audit_handler)
    return audit_logger


audit_logger = configure_project_logging()

for bundled_dir, runtime_dir in [
    (os.path.join(BASE_DIR, "static", "uploads"), UPLOAD_DIR),
    (os.path.join(BASE_DIR, "static", "assets"), ASSET_DIR),
]:
    if bundled_dir == runtime_dir or not os.path.isdir(bundled_dir):
        continue
    for filename in os.listdir(bundled_dir):
        src = os.path.join(bundled_dir, filename)
        dst = os.path.join(runtime_dir, filename)
        if os.path.isfile(src) and not os.path.exists(dst):
            with open(src, "rb") as src_file, open(dst, "wb") as dst_file:
                dst_file.write(src_file.read())

for bundled_file, runtime_file in [
    (os.path.join(BASE_DIR, "records.json"), DATA_FILE),
    (os.path.join(BASE_DIR, "certificates.json"), CERTIFICATE_DATA_FILE),
    (os.path.join(BASE_DIR, "settings.json"), SETTINGS_FILE),
]:
    if bundled_file == runtime_file or not os.path.isfile(bundled_file) or os.path.exists(runtime_file):
        continue
    with open(bundled_file, "rb") as src_file, open(runtime_file, "wb") as dst_file:
        dst_file.write(src_file.read())


def get_client_ip():
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or ""


def read_recent_log_entries(path, limit=200):
    max_lines = max(1, min(int(limit or 200), 1000))
    if not os.path.exists(path):
        return []
    lines = deque(maxlen=max_lines)
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as file_obj:
            for line in file_obj:
                cleaned = line.strip()
                if cleaned:
                    lines.append(cleaned)
    except Exception:
        return []
    return list(lines)


def build_day_wise_audit(entries, per_day=20):
    grouped = {}
    daily_limit = max(1, min(int(per_day or 20), 100))
    for entry in entries:
        time_value = str(entry.get("time", ""))
        day_key = time_value[:10] if len(time_value) >= 10 else "unknown"
        bucket = grouped.setdefault(day_key, {"date": day_key, "count": 0, "recent": []})
        bucket["count"] += 1
        if len(bucket["recent"]) < daily_limit:
            bucket["recent"].append(entry)
    ordered_days = sorted(grouped.keys(), reverse=True)
    return [grouped[day] for day in ordered_days]


def parse_audit_lines(lines):
    parsed = []
    for line in reversed(lines):
        try:
            parsed.append(json.loads(line))
        except Exception:
            parsed.append({"raw": line})
    return parsed


def parse_iso_utc(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    if raw.endswith("Z"):
        raw = raw[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(raw)
    except Exception:
        return None
    if parsed.tzinfo is not None:
        try:
            parsed = parsed.astimezone(timezone.utc).replace(tzinfo=None)
        except Exception:
            parsed = parsed.replace(tzinfo=None)
    return parsed


def filter_audit_entries_since_days(entries, since_days):
    days = max(1, min(int(since_days or 3), 30))
    cutoff_date = (datetime.utcnow() - timedelta(days=days - 1)).date()
    filtered = []
    for entry in entries:
        dt = parse_iso_utc(entry.get("time"))
        if dt and dt.date() >= cutoff_date:
            filtered.append(entry)
    return filtered


def filter_app_entries_since_days(lines, since_days):
    days = max(1, min(int(since_days or 3), 30))
    cutoff_date = (datetime.utcnow() - timedelta(days=days - 1)).date()
    filtered = []
    for line in lines:
        date_part = str(line or "")[:10]
        try:
            day = datetime.strptime(date_part, "%Y-%m-%d").date()
        except Exception:
            continue
        if day >= cutoff_date:
            filtered.append(line)
    return filtered


def filter_audit_entries_for_day(entries, day_key):
    target = str(day_key or "")
    return [entry for entry in entries if str(entry.get("time", ""))[:10] == target]


def filter_app_entries_for_day(lines, day_key):
    target = str(day_key or "")
    return [line for line in lines if str(line or "")[:10] == target]


def day_keys_for_range(since_days):
    days = max(1, min(int(since_days or 3), 30))
    today = datetime.utcnow().date()
    keys = []
    for index in range(days):
        keys.append((today - timedelta(days=index)).strftime("%Y-%m-%d"))
    return keys


def build_supabase_storage_object_url(object_path, bucket_name=None):
    bucket = (bucket_name or app.config["SUPABASE_PHOTOS_BUCKET"]).strip() or app.config["SUPABASE_PHOTOS_BUCKET"]
    object_path = (object_path or "").strip("/")
    base = app.config["SUPABASE_URL"].rstrip("/")
    return f"{base}/storage/v1/object/{urllib_parse.quote(bucket, safe='')}/{urllib_parse.quote(object_path, safe='/')}"


def download_supabase_storage_object_text(object_path, bucket_name=None):
    if not is_supabase_enabled():
        raise RuntimeError("Supabase is not configured")
    url = build_supabase_storage_object_url(object_path, bucket_name)
    req = urllib_request.Request(url, method="GET")
    for key, value in supabase_headers().items():
        if key.lower() == "content-type":
            continue
        req.add_header(key, value)
    with urllib_request.urlopen(req, timeout=60) as response:
        return response.read().decode("utf-8", errors="replace")


def sync_readable_logs_to_supabase(since_days=3, limit=5000, per_day=50):
    if not is_supabase_enabled():
        raise RuntimeError("Supabase is not configured")

    limit_value = max(1, min(int(limit or 5000), 20000))
    per_day_value = max(1, min(int(per_day or 50), 300))
    day_keys = day_keys_for_range(since_days)
    audit_entries = parse_audit_lines(read_recent_log_entries(AUDIT_LOG_FILE, limit_value))
    app_entries = read_recent_log_entries(APP_LOG_FILE, limit_value)

    uploaded_days = []
    total_audit = 0
    total_app = 0
    for day_key in day_keys:
        day_audit = filter_audit_entries_for_day(audit_entries, day_key)
        day_app = filter_app_entries_for_day(app_entries, day_key)
        day_summary = {
            "date": day_key,
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "audit_count": len(day_audit),
            "app_count": len(day_app),
            "day_wise_audit": build_day_wise_audit(day_audit, per_day=per_day_value),
        }
        prefix = f"system-logs/readable/{day_key}"
        upload_bytes_to_supabase_storage(
            "\n".join(json.dumps(item, ensure_ascii=False) for item in day_audit).encode("utf-8"),
            f"{prefix}/audit.jsonl",
            "application/json",
        )
        upload_bytes_to_supabase_storage(
            "\n".join(day_app).encode("utf-8"),
            f"{prefix}/app.log",
            "text/plain",
        )
        upload_bytes_to_supabase_storage(
            json.dumps(day_summary, ensure_ascii=False, indent=2).encode("utf-8"),
            f"{prefix}/summary.json",
            "application/json",
        )
        uploaded_days.append(day_key)
        total_audit += len(day_audit)
        total_app += len(day_app)

    return {
        "uploaded_days": uploaded_days,
        "since_days": len(uploaded_days),
        "total_audit_events": total_audit,
        "total_app_lines": total_app,
    }


def load_readable_logs_from_supabase(since_days=3, per_day=20, limit=1000):
    if not is_supabase_enabled():
        raise RuntimeError("Supabase is not configured")
    limit_value = max(1, min(int(limit or 1000), 5000))
    per_day_value = max(1, min(int(per_day or 20), 100))
    selected_days = set(day_keys_for_range(since_days))
    audit_entries = []
    app_entries = []
    for day_key in sorted(selected_days, reverse=True):
        prefix = f"system-logs/readable/{day_key}"
        try:
            audit_text = download_supabase_storage_object_text(f"{prefix}/audit.jsonl")
            for line in audit_text.splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    audit_entries.append(json.loads(line))
                except Exception:
                    continue
        except Exception:
            pass
        try:
            app_text = download_supabase_storage_object_text(f"{prefix}/app.log")
            for line in app_text.splitlines():
                cleaned = line.strip()
                if cleaned:
                    app_entries.append(cleaned)
        except Exception:
            pass

    audit_entries = audit_entries[:limit_value]
    app_entries = app_entries[:limit_value]
    return {
        "audit_log": audit_entries,
        "day_wise_audit": build_day_wise_audit(audit_entries, per_day=per_day_value),
        "app_log": app_entries,
        "source": "supabase_readable",
    }


def build_activity_logs_archive(limit=1000, per_day=25, since_days=3):
    limit_value = max(1, min(int(limit or 1000), 2000))
    per_day_value = max(1, min(int(per_day or 25), 200))
    since_days_value = max(1, min(int(since_days or 3), 30))
    audit_entries = filter_audit_entries_since_days(
        parse_audit_lines(read_recent_log_entries(AUDIT_LOG_FILE, limit_value)),
        since_days_value,
    )
    app_entries = filter_app_entries_since_days(read_recent_log_entries(APP_LOG_FILE, limit_value), since_days_value)
    day_wise = build_day_wise_audit(audit_entries, per_day=per_day_value)
    summary = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "limit": limit_value,
        "per_day": per_day_value,
        "since_days": since_days_value,
        "total_audit_events": len(audit_entries),
        "total_app_lines": len(app_entries),
        "days": [
            {"date": item.get("date", ""), "count": item.get("count", 0)}
            for item in day_wise
        ],
        "day_wise_audit": day_wise,
    }

    archive_buffer = io.BytesIO()
    with zipfile.ZipFile(archive_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("activity_summary.json", json.dumps(summary, indent=2, ensure_ascii=False))
        zf.writestr(
            "audit_recent.jsonl",
            "\n".join(json.dumps(item, ensure_ascii=False) for item in audit_entries),
        )
        zf.writestr("app_recent.log", "\n".join(app_entries))
    return archive_buffer.getvalue(), summary


def log_audit_event(event_type, **extra):
    payload = {
        "time": datetime.utcnow().isoformat() + "Z",
        "event": event_type,
        "request_id": getattr(g, "request_id", ""),
    }
    payload.update(extra)
    try:
        audit_logger.info(json.dumps(payload, ensure_ascii=False, default=str))
    except Exception:
        app.logger.warning("Failed to write audit event", exc_info=True)


@app.before_request
def begin_request_trace():
    g.request_started_at = time.perf_counter()
    g.request_id = uuid.uuid4().hex[:12]


@app.after_request
def end_request_trace(response):
    started = getattr(g, "request_started_at", None)
    duration_ms = round((time.perf_counter() - started) * 1000, 2) if started is not None else None
    log_audit_event(
        "http_request",
        method=request.method,
        path=request.path,
        query=request.query_string.decode("utf-8", errors="ignore"),
        status=response.status_code,
        duration_ms=duration_ms,
        ip=get_client_ip(),
        user_agent=request.headers.get("User-Agent", ""),
        admin_authenticated=bool(session.get("admin_authenticated")),
    )
    response.headers["X-Request-Id"] = g.request_id
    return response


@app.teardown_request
def trace_request_exception(exc):
    if not exc:
        return
    log_audit_event(
        "request_exception",
        method=request.method,
        path=request.path,
        error=str(exc),
        ip=get_client_ip(),
    )

def default_institute_settings(institute=None):
    return {
        "institute_name": institute or "",
        "background_url": "",
        "background_drive_id": "",
        "print_background_active_id": "",
        "print_backgrounds": [],
        "office_backgrounds": [],
        "fabric_batch_background_bindings": {},
        "certificate_background_url": "",
        "certificate_background_drive_id": "",
        "signature_url": "",
        "signature_drive_id": "",
        "print_templates": [],
        "facility_custom_sub_locations": {},
    }


def default_admin_prefs():
    return {
        "storage_quota_mb": 1024,
        "auto_delete_enabled": False,
        "auto_delete_days": 15,
        "last_auto_cleanup_at": "",
        "fabric_global_backgrounds": [],
    }


FACILITY_LOCATION_INSTITUTES = [
    "Govt. Community Health Centre (CHC), Jhunjhunu",
    "Govt. Primary Health Centre (PHC), Jhunjhunu",
    "Govt. Health Sub Centre, Jhunjhunu",
]

FACILITY_LOCATIONS = {
    "Govt. Community Health Centre (CHC), Jhunjhunu": ["Jhunjhunu", "Chirawa", "Khetri", "Udaipurwati", "Nawalgarh", "Buhana", "Malsisar", "Mandawa", "Surajgarh"],
    "Govt. Primary Health Centre (PHC), Jhunjhunu": ["All PHCs", "Jhunjhunu", "Chirawa", "Surajgarh", "Khetri", "Buhana", "Nawalgarh", "Udaipurwati", "Malsisar"],
    "Govt. Health Sub Centre, Jhunjhunu": ["All Sub Centres", "Jhunjhunu", "Chirawa", "Buhana", "Khetri", "Nawalgarh", "Udaipurwati", "Pilani", "Surajgarh", "Mandawa"],
}

FACILITY_SUB_LOCATIONS = {
    "Govt. Community Health Centre (CHC), Jhunjhunu": {
        "Jhunjhunu": ["Bagar", "Indali", "Baragaon"],
        "Chirawa": ["Chirawa", "Mandrella"],
        "Khetri": ["Khetri", "Singhana"],
        "Udaipurwati": ["Udaipurwati", "Gudhagorji", "Chirana", "Ponkh"],
        "Nawalgarh": ["Nawalgarh", "Khirod", "Parasrampura"],
        "Buhana": ["Buhana"],
        "Malsisar": ["Malsisar", "Bissau", "Mahansar"],
        "Mandawa": ["Mandawa"],
        "Surajgarh": ["Surajgarh"],
    },
    "Govt. Primary Health Centre (PHC), Jhunjhunu": {
        "All PHCs": [
            "Islampur", "Kali Pahari", "Binjusar", "Hetamsar", "Bharu", "Bhadarwas", "Bharunda Kalan", "Patusari", "Churi", "Nua",
            "Bakhtawarpura", "Padampura", "Solana", "Sultana", "Chanana", "Jakhora", "Ardawata", "Bangothri Kalan", "Kakoda", "Kajara",
            "Kidwana", "Pipli", "Chhapara", "Devroad", "Jakhod", "Babai", "Shimla", "Sihod", "Papurana", "Rasulpur", "Tyonda", "Tatija",
            "Jasarapur", "Tibba Basai", "Udamandi", "Kuhadwas", "Singhana", "Hirwa", "Pacheri Kalan", "Sawlod", "Sohali", "Churina",
            "Jhajhar", "Basawa", "Mandasi", "Khirod", "Dhigal", "Bhagera", "Parasrampura", "Jakhal", "Jejusar", "Gothara", "Baragaun",
            "Mandawara", "Bhorki", "Ponkh", "Pachlangi", "Chhapoli", "Chanwara", "Titanwad", "Gudha Gorji", "Alsisar", "Birmi",
            "Gangiyasar", "Kaant", "Kaliyasar", "Ladusar", "Mahansar", "Niradhanu", "Tumkor",
        ],
        "Jhunjhunu": ["Islampur", "Bagar", "Hetamsar", "Indali", "Bharu", "Sultana", "Dhigal", "Birmi", "Bajla", "Gura", "Titanwad"],
        "Chirawa": ["Bakhtawarpura", "Padampura", "Solana", "Sultana", "Chanana", "Jakhora", "Ardawata", "Bharunda Kalan"],
        "Surajgarh": ["Bangothri Kalan", "Kakoda", "Kajara", "Kidwana", "Pipli", "Chhapara", "Devroad", "Jakhod"],
        "Khetri": ["Babai", "Shimla", "Sihod", "Papurana", "Rasulpur", "Tyonda", "Tatija", "Jasarapur", "Tibba Basai"],
        "Buhana": ["Udamandi", "Kuhadwas", "Singhana", "Hirwa", "Pacheri Kalan", "Sawlod", "Sohali", "Churina"],
        "Nawalgarh": ["Jhajhar", "Basawa", "Mandasi", "Khirod", "Dhigal", "Bhagera", "Parasrampura", "Jakhal", "Jejusar", "Gothara"],
        "Udaipurwati": ["Baragaun", "Mandawara", "Bhorki", "Ponkh", "Pachlangi", "Chhapoli", "Chanwara", "Titanwad", "Gudha Gorji"],
        "Malsisar": ["Alsisar", "Birmi", "Gangiyasar", "Kaant", "Kaliyasar", "Ladusar", "Mahansar", "Niradhanu", "Tumkor"],
    },
    "Govt. Health Sub Centre, Jhunjhunu": {
        "All Sub Centres": [
            "Jhunjhunu Town", "Badalwas", "Dundlod", "Sultana", "Mandrella", "Nua", "Khudana", "Chirawa Town", "Mandrela", "Pilod",
            "Arnia", "Kithana", "Bajawa", "Gidania", "Buhana Town", "Pacheri Kalan", "Sohli", "Singhana", "Bhirr", "Kuharwas", "Bhukana",
            "Khetri Town", "Kolihan Nagar", "Banwas", "Ravla", "Papurna", "Ajeetpura", "Shimla", "Nawalgarh Town", "Mukasundgarh",
            "Parsurampura", "Nayi Basti", "Jakhora", "Barwasi", "Basawa", "Udaipurwati Town", "Gudha Gorji", "Posana", "Chhapoli",
            "Bagora", "Mankwas", "Morda", "Pilani Town", "Chidawa Road", "Hanutpura", "Basant Vihar", "Dulania", "Morwa", "Khedla",
            "Surajgarh Town", "Kajra", "Loharu Road", "Malsisar", "Bhojasar", "Kakoda", "Jakhod", "Mandawa Town", "Alsisar", "Baloda",
            "Bissau", "Tain", "Sehkalon Ki Dhani", "Hanumanpura",
        ],
        "Jhunjhunu": ["Jhunjhunu Town", "Badalwas", "Dundlod", "Sultana", "Mandrella", "Nua", "Khudana"],
        "Chirawa": ["Chirawa Town", "Mandrela", "Pilod", "Arnia", "Kithana", "Bajawa", "Gidania"],
        "Buhana": ["Buhana Town", "Pacheri Kalan", "Sohli", "Singhana", "Bhirr", "Kuharwas", "Bhukana"],
        "Khetri": ["Khetri Town", "Kolihan Nagar", "Banwas", "Ravla", "Papurna", "Ajeetpura", "Shimla"],
        "Nawalgarh": ["Nawalgarh Town", "Mukasundgarh", "Parsurampura", "Nayi Basti", "Jakhora", "Barwasi", "Basawa"],
        "Udaipurwati": ["Udaipurwati Town", "Gudha Gorji", "Posana", "Chhapoli", "Bagora", "Mankwas", "Morda"],
        "Pilani": ["Pilani Town", "Chidawa Road", "Hanutpura", "Basant Vihar", "Dulania", "Morwa", "Khedla"],
        "Surajgarh": ["Surajgarh Town", "Kajra", "Loharu Road", "Malsisar", "Bhojasar", "Kakoda", "Jakhod"],
        "Mandawa": ["Mandawa Town", "Alsisar", "Baloda", "Bissau", "Tain", "Sehkalon Ki Dhani", "Hanumanpura"],
    },
}


def make_storage_slug(value):
    safe = secure_filename((value or "").strip()).lower()
    return safe or "default"


def make_storage_filename(kind, institute):
    return f"{kind}__{make_storage_slug(institute)}.json"


def make_storage_path(kind, institute):
    directory_map = {
        "records": RECORDS_DIR,
        "certificates": CERTIFICATES_DIR,
        "settings": SETTINGS_DIR,
    }
    return os.path.join(directory_map[kind], make_storage_filename(kind, institute))


def list_local_storage_filenames(kind):
    directory_map = {
        "records": RECORDS_DIR,
        "certificates": CERTIFICATES_DIR,
        "settings": SETTINGS_DIR,
    }
    directory = directory_map[kind]
    if not os.path.isdir(directory):
        return []
    return sorted(name for name in os.listdir(directory) if name.startswith(f"{kind}__") and name.endswith(".json"))


def list_drive_storage_filenames(kind):
    return []


def load_legacy_records():
    return load_json_store(DATA_FILE, "records.json", [])


def load_legacy_certificates():
    return load_json_store(CERTIFICATE_DATA_FILE, "certificates.json", [])


def load_legacy_settings():
    defaults = default_institute_settings()
    defaults.update({
        "backgrounds": {},
        "background_drive_ids": {},
        "certificate_backgrounds": {},
        "certificate_background_drive_ids": {},
    })
    loaded = load_json_store(SETTINGS_FILE, "settings.json", defaults)
    if isinstance(loaded, dict):
        defaults.update(loaded)
    return defaults


def load_admin_prefs():
    defaults = default_admin_prefs()
    loaded = load_json_store(ADMIN_PREFS_FILE, "admin_prefs.json", defaults)
    if isinstance(loaded, dict):
        defaults.update(loaded)
    try:
        defaults["storage_quota_mb"] = max(1, int(defaults.get("storage_quota_mb", 1024) or 1024))
    except Exception:
        defaults["storage_quota_mb"] = 1024
    try:
        defaults["auto_delete_days"] = max(1, int(defaults.get("auto_delete_days", 15) or 15))
    except Exception:
        defaults["auto_delete_days"] = 15
    defaults["auto_delete_enabled"] = bool(defaults.get("auto_delete_enabled"))
    defaults["last_auto_cleanup_at"] = str(defaults.get("last_auto_cleanup_at") or "")
    defaults["fabric_global_backgrounds"] = sanitize_global_backgrounds_list(defaults.get("fabric_global_backgrounds", []))
    return defaults


def save_admin_prefs(prefs):
    payload = default_admin_prefs()
    if isinstance(prefs, dict):
        payload.update(prefs)
    payload["storage_quota_mb"] = max(1, int(payload.get("storage_quota_mb", 1024) or 1024))
    payload["auto_delete_days"] = max(1, int(payload.get("auto_delete_days", 15) or 15))
    payload["auto_delete_enabled"] = bool(payload.get("auto_delete_enabled"))
    payload["last_auto_cleanup_at"] = str(payload.get("last_auto_cleanup_at") or "")
    payload["fabric_global_backgrounds"] = sanitize_global_backgrounds_list(payload.get("fabric_global_backgrounds", []))
    save_json_store(ADMIN_PREFS_FILE, "admin_prefs.json", payload)


def migrate_legacy_records_if_needed():
    legacy_records = load_legacy_records()
    if not legacy_records:
        return
    existing = set(list_local_storage_filenames("records"))
    grouped = {}
    for record in legacy_records:
        institute = canonicalize_institute_name(record.get("institute_name"))
        if not institute:
            continue
        grouped.setdefault(institute, []).append(record)
    for institute, records in grouped.items():
        filename = make_storage_filename("records", institute)
        if filename not in existing:
            save_json_store(make_storage_path("records", institute), filename, records)
            existing.add(filename)


def migrate_legacy_certificates_if_needed():
    legacy_certificates = load_legacy_certificates()
    if not legacy_certificates:
        return
    existing = set(list_local_storage_filenames("certificates"))
    grouped = {}
    for certificate in legacy_certificates:
        institute = canonicalize_institute_name(certificate.get("institute_name"))
        if not institute:
            continue
        grouped.setdefault(institute, []).append(certificate)
    for institute, certificates in grouped.items():
        filename = make_storage_filename("certificates", institute)
        if filename not in existing:
            save_json_store(make_storage_path("certificates", institute), filename, certificates)
            existing.add(filename)


def migrate_legacy_settings_if_needed(institute):
    if not institute:
        return
    path = make_storage_path("settings", institute)
    filename = make_storage_filename("settings", institute)
    if os.path.exists(path):
        return
    legacy = load_legacy_settings()
    settings = default_institute_settings(institute)
    settings["background_url"] = legacy.get("backgrounds", {}).get(institute, "") or legacy.get("background_url", "")
    settings["background_drive_id"] = legacy.get("background_drive_ids", {}).get(institute, "")
    settings["certificate_background_url"] = legacy.get("certificate_backgrounds", {}).get(institute, "")
    settings["certificate_background_drive_id"] = legacy.get("certificate_background_drive_ids", {}).get(institute, "")
    settings["signature_url"] = legacy.get("signature_url", "")
    settings["signature_drive_id"] = legacy.get("signature_drive_id", "")
    if any(settings.get(key) for key in settings if key != "institute_name"):
        save_json_store(path, filename, settings)


def load_records(institute=None):
    migrate_legacy_records_if_needed()
    institute = canonicalize_institute_name(institute)
    if institute:
        return load_json_store(make_storage_path("records", institute), make_storage_filename("records", institute), [])

    records = []
    filenames = sorted(set(list_local_storage_filenames("records")))
    for filename in filenames:
        records.extend(load_json_store(os.path.join(RECORDS_DIR, filename), filename, []))
    return records


def save_records(records, institute):
    institute = canonicalize_institute_name(institute)
    if not institute:
        raise ValueError("Institute is required to save records")
    save_json_store(make_storage_path("records", institute), make_storage_filename("records", institute), records)


def load_certificates(institute=None):
    migrate_legacy_certificates_if_needed()
    institute = canonicalize_institute_name(institute)
    if institute:
        return load_json_store(make_storage_path("certificates", institute), make_storage_filename("certificates", institute), [])

    certificates = []
    filenames = sorted(set(list_local_storage_filenames("certificates")))
    for filename in filenames:
        certificates.extend(load_json_store(os.path.join(CERTIFICATES_DIR, filename), filename, []))
    return certificates


def save_certificates(certificates, institute):
    institute = canonicalize_institute_name(institute)
    if not institute:
        raise ValueError("Institute is required to save certificates")
    save_json_store(make_storage_path("certificates", institute), make_storage_filename("certificates", institute), certificates)


def load_settings(institute=None):
    institute = canonicalize_institute_name(institute)
    defaults = default_institute_settings(institute)
    if not institute:
        return defaults
    migrate_legacy_settings_if_needed(institute)
    loaded = load_json_store(make_storage_path("settings", institute), make_storage_filename("settings", institute), defaults)
    if isinstance(loaded, dict):
        defaults.update(loaded)
    defaults["institute_name"] = institute
    ensure_print_background_state(defaults)
    defaults["office_backgrounds"] = sanitize_office_backgrounds_list(defaults.get("office_backgrounds", []))
    return defaults


def save_settings(settings, institute):
    institute = canonicalize_institute_name(institute or settings.get("institute_name"))
    if not institute:
        raise ValueError("Institute is required to save settings")
    payload = default_institute_settings(institute)
    if isinstance(settings, dict):
        payload.update(settings)
    payload["institute_name"] = institute
    ensure_print_background_state(payload)
    payload["office_backgrounds"] = sanitize_office_backgrounds_list(payload.get("office_backgrounds", []))
    save_json_store(make_storage_path("settings", institute), make_storage_filename("settings", institute), payload)


def sanitize_facility_custom_sub_locations(data):
    cleaned = {}
    if not isinstance(data, dict):
        return cleaned
    for block, values in data.items():
        block_name = str(block or "").strip()
        if not block_name or not isinstance(values, list):
            continue
        items = []
        for value in values:
            text = str(value or "").strip()
            if text and text not in items and text != "Add New":
                items.append(text)
        if items:
            cleaned[block_name] = sorted(items)
    return cleaned


def build_facility_structure_payload():
    structure = {}
    for institute in FACILITY_LOCATION_INSTITUTES:
        settings = load_settings(institute)
        custom_map = sanitize_facility_custom_sub_locations(settings.get("facility_custom_sub_locations", {}))
        merged_sub_locations = {}
        for block in FACILITY_LOCATIONS.get(institute, []):
            base_items = list(FACILITY_SUB_LOCATIONS.get(institute, {}).get(block, []))
            custom_items = list(custom_map.get(block, []))
            merged = []
            for item in base_items + custom_items:
                text = str(item or "").strip()
                if text and text != "Add New" and text not in merged:
                    merged.append(text)
            merged_sub_locations[block] = merged
        structure[institute] = {
            "blocks": list(FACILITY_LOCATIONS.get(institute, [])),
            "sub_locations": merged_sub_locations,
        }
    return structure


def clamp_int(value, fallback, minimum, maximum):
    try:
        parsed = int(float(value))
    except Exception:
        return fallback
    return max(minimum, min(maximum, parsed))


def clamp_float(value, fallback, minimum, maximum):
    try:
        parsed = float(value)
    except Exception:
        return fallback
    return max(minimum, min(maximum, parsed))


def normalize_hex_color(value, fallback):
    color = str(value or "").strip()
    if HEX_COLOR_PATTERN.match(color):
        return color.lower()
    return fallback


def normalize_foot_bg(value, fallback):
    text = str(value or "").strip()
    if not text:
        return fallback
    if len(text) > 80:
        return fallback
    if text.lower().startswith("rgba(") or text.lower().startswith("rgb("):
        return text
    return fallback


def sanitize_print_template_config(config):
    if not isinstance(config, dict):
        return None
    raw_transforms = config.get("field_transforms", {}) if isinstance(config.get("field_transforms", {}), dict) else {}
    sanitized_transforms = {}
    for field_id, value in raw_transforms.items():
        field_key = str(field_id or "").strip()
        if not field_key or not isinstance(value, dict):
            continue
        sanitized_transforms[field_key] = {
            "x": clamp_int(value.get("x"), 0, -80, 80),
            "y": clamp_int(value.get("y"), 0, -80, 80),
            "scale": clamp_int(value.get("scale"), 100, 60, 160),
        }
    return {
        "accent": normalize_hex_color(config.get("accent"), "#8b0000"),
        "head_end": normalize_hex_color(config.get("head_end"), "#5d0000"),
        "role": normalize_hex_color(config.get("role"), "#9a6a18"),
        "gap": clamp_int(config.get("gap"), 12, 6, 24),
        "radius": clamp_int(config.get("radius"), 14, 4, 28),
        "opacity": round(clamp_float(config.get("opacity"), 0.14, 0, 0.30), 2),
        "name": clamp_int(config.get("name"), 28, 22, 34),
        "border": normalize_hex_color(config.get("border"), "#d8c9bd"),
        "foot": normalize_foot_bg(config.get("foot"), "rgba(255,248,239,0.92)"),
        "badge": normalize_hex_color(config.get("badge"), "#8b0000"),
        "photo": clamp_int(config.get("photo"), 106, 90, 130),
        "field_transforms": sanitized_transforms,
    }


def sanitize_print_templates_list(items):
    if not isinstance(items, list):
        return []
    sanitized = []
    seen_ids = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        template_id = str(item.get("id") or "").strip()
        name = str(item.get("name") or "").strip()
        config = sanitize_print_template_config(item.get("config"))
        if not template_id or not name or not config or template_id in seen_ids:
            continue
        seen_ids.add(template_id)
        sanitized.append({
            "id": template_id,
            "name": name[:80],
            "config": config,
            "created_at": str(item.get("created_at") or ""),
            "updated_at": str(item.get("updated_at") or ""),
        })
    return sanitized[:60]


def sanitize_print_backgrounds_list(items):
    if not isinstance(items, list):
        return []
    sanitized = []
    seen_ids = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        background_id = str(item.get("id") or "").strip()
        name = str(item.get("name") or "").strip()
        url = str(item.get("url") or "").strip()
        if not background_id or not url or background_id in seen_ids:
            continue
        seen_ids.add(background_id)
        orientation = str(item.get("orientation") or "landscape").strip().lower()
        if orientation not in ("landscape", "portrait"):
            orientation = "landscape"
        sanitized.append({
            "id": background_id,
            "name": name[:80] if name else "Background",
            "url": url,
            "orientation": orientation,
            "created_at": str(item.get("created_at") or ""),
            "updated_at": str(item.get("updated_at") or ""),
        })
    return sanitized[:80]


def sanitize_office_backgrounds_list(items):
    if not isinstance(items, list):
        return []
    sanitized = []
    seen_ids = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        background_id = str(item.get("id") or "").strip()
        name = str(item.get("name") or "").strip()
        url = str(item.get("url") or "").strip()
        block = str(item.get("block") or "").strip()
        facility_sub_location = str(item.get("facility_sub_location") or "").strip()
        if not background_id or not url or background_id in seen_ids or not block or not facility_sub_location:
            continue
        seen_ids.add(background_id)
        orientation = str(item.get("orientation") or "landscape").strip().lower()
        if orientation not in ("landscape", "portrait"):
            orientation = "landscape"
        sanitized.append({
            "id": background_id,
            "name": name[:80] if name else "Office Background",
            "url": url,
            "block": block,
            "facility_sub_location": facility_sub_location,
            "orientation": orientation,
            "created_at": str(item.get("created_at") or ""),
            "updated_at": str(item.get("updated_at") or ""),
        })
    return sanitized[:200]


def normalize_url_lookup_key(value):
    return str(value or "").strip().split("?", 1)[0].lower()


def sanitize_global_backgrounds_list(items):
    if not isinstance(items, list):
        return []
    sanitized = []
    seen_ids = set()
    seen_urls = set()
    for item in items:
        if not isinstance(item, dict):
            continue
        background_id = str(item.get("id") or "").strip()
        url = str(item.get("url") or "").strip()
        key = normalize_url_lookup_key(url)
        if not background_id or not url or background_id in seen_ids or key in seen_urls:
            continue
        seen_ids.add(background_id)
        seen_urls.add(key)
        orientation = str(item.get("orientation") or "landscape").strip().lower()
        if orientation not in ("landscape", "portrait"):
            orientation = "landscape"
        sanitized.append({
            "id": background_id,
            "name": str(item.get("name") or "Background")[:80],
            "url": url,
            "orientation": orientation,
            "created_at": str(item.get("created_at") or ""),
            "updated_at": str(item.get("updated_at") or ""),
            "institute_name": canonicalize_institute_name(item.get("institute_name")) or "",
        })
    return sanitized[:2000]


def build_print_background_filename(institute_name, background_id):
    return f"print_bg_{make_asset_slug(institute_name)}_{secure_filename(background_id)}.jpg"


def ensure_print_background_state(settings):
    backgrounds = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))
    active_id = str(settings.get("print_background_active_id") or "").strip()
    active_item = next((item for item in backgrounds if item.get("id") == active_id), None)

    if not active_item and settings.get("background_url"):
        fallback_id = f"legacy-{make_storage_slug(settings.get('institute_name'))}"
        fallback_item = {
            "id": fallback_id,
            "name": "Primary Background",
            "url": settings.get("background_url"),
            "orientation": "landscape",
            "created_at": "",
            "updated_at": "",
        }
        backgrounds.insert(0, fallback_item)
        active_item = fallback_item
        active_id = fallback_id

    if active_item:
        settings["background_url"] = active_item.get("url", "")
    elif backgrounds:
        settings["background_url"] = backgrounds[0].get("url", "")
        active_id = backgrounds[0].get("id", "")
    else:
        settings["background_url"] = ""
        active_id = ""

    settings["print_backgrounds"] = backgrounds
    settings["print_background_active_id"] = active_id
    return settings


def sanitize_fabric_batch_background_bindings(value):
    if not isinstance(value, dict):
        return {}
    cleaned = {}
    for key, bg_url in value.items():
        batch_id = str(key or "").strip()
        url_value = str(bg_url or "").strip()
        if not batch_id:
            continue
        cleaned[batch_id] = url_value
    return cleaned


def list_known_institutes_from_settings():
    institutes = set()
    if not os.path.isdir(SETTINGS_DIR):
        return institutes
    for filename in os.listdir(SETTINGS_DIR):
        if not (filename.startswith("settings__") and filename.endswith(".json")):
            continue
        file_path = os.path.join(SETTINGS_DIR, filename)
        payload = load_json_store(file_path, filename, {})
        if not isinstance(payload, dict):
            continue
        institute = canonicalize_institute_name(payload.get("institute_name"))
        if institute:
            institutes.add(institute)
    return institutes


def remove_background_url_from_design_payload(design_payload, target_url_key):
    if not isinstance(design_payload, dict) or not target_url_key:
        return False
    changed = False

    def clear_snapshot(snapshot):
        nonlocal changed
        if not isinstance(snapshot, dict):
            return
        asset_urls = snapshot.get("asset_urls")
        if isinstance(asset_urls, dict):
            bg_url = str(asset_urls.get("background_image") or "").strip()
            if normalize_url_lookup_key(bg_url) == target_url_key:
                asset_urls.pop("background_image", None)
                changed = True
        library = snapshot.get("background_library")
        if isinstance(library, list):
            filtered = [item for item in library if normalize_url_lookup_key(item) != target_url_key]
            if len(filtered) != len(library):
                snapshot["background_library"] = filtered
                changed = True

    clear_snapshot(design_payload)
    side_designs = design_payload.get("side_designs")
    if isinstance(side_designs, dict):
        clear_snapshot(side_designs.get("front"))
        clear_snapshot(side_designs.get("back"))
    return changed


def remove_global_background_from_all_settings(target_url):
    target_key = normalize_url_lookup_key(target_url)
    if not target_key:
        return {"updated_institutes": 0, "removed_background_links": 0, "removed_batch_bindings": 0, "updated_designs": 0}

    institutes = set(list_known_institutes_from_settings())
    try:
        for batch in list_all_supabase_batches():
            institute_name = canonicalize_institute_name(batch.get("institute_name"))
            if institute_name:
                institutes.add(institute_name)
    except Exception:
        pass

    updated_institutes = 0
    removed_links = 0
    removed_bindings = 0
    updated_designs = 0

    for institute in sorted(institutes):
        settings = load_settings(institute)
        institute_changed = False

        backgrounds = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))
        filtered_backgrounds = [item for item in backgrounds if normalize_url_lookup_key(item.get("url")) != target_key]
        if len(filtered_backgrounds) != len(backgrounds):
            removed_links += len(backgrounds) - len(filtered_backgrounds)
            settings["print_backgrounds"] = filtered_backgrounds
            institute_changed = True

        bindings = sanitize_fabric_batch_background_bindings(settings.get("fabric_batch_background_bindings", {}))
        filtered_bindings = {
            batch_id: bg_url for batch_id, bg_url in bindings.items()
            if normalize_url_lookup_key(bg_url) != target_key
        }
        if len(filtered_bindings) != len(bindings):
            removed_bindings += len(bindings) - len(filtered_bindings)
            settings["fabric_batch_background_bindings"] = filtered_bindings
            institute_changed = True

        background_url = str(settings.get("background_url") or "").strip()
        if normalize_url_lookup_key(background_url) == target_key:
            settings["background_url"] = ""
            settings["background_drive_id"] = ""
            institute_changed = True

        design_payload = settings.get("fabric_design")
        if remove_background_url_from_design_payload(design_payload, target_key):
            settings["fabric_design"] = design_payload
            updated_designs += 1
            institute_changed = True

        ensure_print_background_state(settings)
        if institute_changed:
            save_settings(settings, institute)
            updated_institutes += 1

    return {
        "updated_institutes": updated_institutes,
        "removed_background_links": removed_links,
        "removed_batch_bindings": removed_bindings,
        "updated_designs": updated_designs,
    }


def is_supabase_enabled():
    return bool(app.config["SUPABASE_URL"] and app.config["SUPABASE_SERVICE_ROLE_KEY"])


def supabase_headers():
    key = app.config["SUPABASE_SERVICE_ROLE_KEY"]
    return {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }


def supabase_storage_headers(content_type=None):
    headers = {
        "apikey": app.config["SUPABASE_SERVICE_ROLE_KEY"],
        "Authorization": f"Bearer {app.config['SUPABASE_SERVICE_ROLE_KEY']}",
        "x-upsert": "true",
    }
    if content_type:
        headers["Content-Type"] = content_type
    return headers


def supabase_request(method, path, payload=None, query=None, prefer_representation=False):
    if not is_supabase_enabled():
        raise RuntimeError("Supabase is not configured")
    base = app.config["SUPABASE_URL"].rstrip("/")
    url = f"{base}/rest/v1/{path.lstrip('/')}"
    if query:
        query_items = []
        for key, value in query.items():
            if value is None:
                continue
            query_items.append((key, value))
        if query_items:
            url += "?" + urllib_parse.urlencode(query_items)
    headers = supabase_headers()
    if prefer_representation:
        headers["Prefer"] = "return=representation"
    data = None
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
    req = urllib_request.Request(url, data=data, method=method.upper())
    for key, value in headers.items():
        req.add_header(key, value)
    try:
        with urllib_request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else []
    except urllib_error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(raw or f"Supabase request failed with status {exc.code}") from exc


def ensure_supabase_bucket(bucket_name):
    if not is_supabase_enabled():
        raise RuntimeError("Supabase is not configured")
    base = app.config["SUPABASE_URL"].rstrip("/")
    url = f"{base}/storage/v1/bucket/{urllib_parse.quote(bucket_name, safe='')}"
    req = urllib_request.Request(url, method="GET")
    for key, value in supabase_headers().items():
        req.add_header(key, value)
    try:
        with urllib_request.urlopen(req, timeout=30):
            return
    except urllib_error.HTTPError as exc:
        if exc.code != 404:
            raw = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(raw or f"Bucket lookup failed with status {exc.code}") from exc

    create_req = urllib_request.Request(
        f"{base}/storage/v1/bucket",
        data=json.dumps({"id": bucket_name, "name": bucket_name, "public": True}).encode("utf-8"),
        method="POST",
    )
    for key, value in supabase_headers().items():
        create_req.add_header(key, value)
    try:
        with urllib_request.urlopen(create_req, timeout=30):
            return
    except urllib_error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        if exc.code == 409:
            return
        raise RuntimeError(raw or f"Bucket create failed with status {exc.code}") from exc


def upload_bytes_to_supabase_storage(file_bytes, object_path, mime_type):
    bucket = app.config["SUPABASE_PHOTOS_BUCKET"]
    ensure_supabase_bucket(bucket)
    base = app.config["SUPABASE_URL"].rstrip("/")
    object_path = object_path.strip("/")
    url = f"{base}/storage/v1/object/{urllib_parse.quote(bucket, safe='')}/{urllib_parse.quote(object_path, safe='/')}"
    public_url = f"{base}/storage/v1/object/public/{urllib_parse.quote(bucket, safe='')}/{urllib_parse.quote(object_path, safe='/')}"

    def attempt_upload():
        req = urllib_request.Request(url, data=file_bytes, method="POST")
        for key, value in supabase_storage_headers(mime_type).items():
            req.add_header(key, value)
        with urllib_request.urlopen(req, timeout=60):
            return public_url

    try:
        return attempt_upload()
    except urllib_error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        if exc.code == 404 and "Bucket not found" in raw:
            ensure_supabase_bucket(bucket)
            try:
                return attempt_upload()
            except urllib_error.HTTPError as retry_exc:
                retry_raw = retry_exc.read().decode("utf-8", errors="replace")
                raise RuntimeError(
                    retry_raw or f'Supabase storage bucket "{bucket}" is missing or unavailable'
                ) from retry_exc
        raise RuntimeError(raw or f"Supabase storage upload failed with status {exc.code}") from exc


def list_supabase_storage_objects(prefix="", limit=1000):
    if not is_supabase_enabled():
        raise RuntimeError("Supabase is not configured")
    bucket = app.config["SUPABASE_PHOTOS_BUCKET"]
    base = app.config["SUPABASE_URL"].rstrip("/")
    url = f"{base}/storage/v1/object/list/{urllib_parse.quote(bucket, safe='')}"
    payload = {
        "prefix": (prefix or "").strip("/"),
        "limit": max(1, min(int(limit or 1000), 1000)),
        "offset": 0,
        "sortBy": {"column": "name", "order": "asc"},
    }
    req = urllib_request.Request(url, data=json.dumps(payload).encode("utf-8"), method="POST")
    headers = supabase_headers()
    headers["Content-Type"] = "application/json"
    for key, value in headers.items():
        req.add_header(key, value)
    try:
        with urllib_request.urlopen(req, timeout=60) as resp:
            raw = resp.read().decode("utf-8")
            data = json.loads(raw) if raw else []
            return data if isinstance(data, list) else []
    except urllib_error.HTTPError as exc:
        raw = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(raw or f"Supabase storage list failed with status {exc.code}") from exc


def delete_supabase_storage_object_path(object_path, bucket_name=None):
    if not is_supabase_enabled():
        return False
    object_path = (object_path or "").strip("/")
    if not object_path:
        return False
    bucket = (bucket_name or app.config["SUPABASE_PHOTOS_BUCKET"]).strip() or app.config["SUPABASE_PHOTOS_BUCKET"]
    base = app.config["SUPABASE_URL"].rstrip("/")
    url = f"{base}/storage/v1/object/{urllib_parse.quote(bucket, safe='')}/{urllib_parse.quote(object_path, safe='/')}"
    req = urllib_request.Request(url, method="DELETE")
    for key, value in supabase_headers().items():
        req.add_header(key, value)
    try:
        with urllib_request.urlopen(req, timeout=30):
            return True
    except urllib_error.HTTPError:
        return False


def list_supabase_storage_file_paths(prefix=""):
    queue = deque([(prefix or "").strip("/")])
    visited = set()
    file_paths = []
    while queue:
        current_prefix = queue.popleft()
        if current_prefix in visited:
            continue
        visited.add(current_prefix)
        items = list_supabase_storage_objects(current_prefix)
        for item in items:
            name = str(item.get("name") or "").strip()
            if not name:
                continue
            metadata = item.get("metadata") or {}
            is_folder = item.get("id") is None and not metadata.get("size")
            if is_folder:
                next_prefix = f"{current_prefix}/{name}".strip("/") if current_prefix else name
                queue.append(next_prefix)
                continue
            object_path = f"{current_prefix}/{name}".strip("/") if current_prefix else name
            file_paths.append(object_path)
    return file_paths


def extract_storage_day_key(object_path):
    value = str(object_path or "").strip()
    if not value:
        return None
    match = re.search(r"(?:^|/)(\d{4}-\d{2}-\d{2})(?:/|$)", value)
    if not match:
        return None
    try:
        return datetime.strptime(match.group(1), "%Y-%m-%d").date()
    except Exception:
        return None


def cleanup_supabase_log_archives(auto_delete_days):
    result = {
        "deleted_log_archives": 0,
        "log_archive_errors": 0,
    }
    if not is_supabase_enabled():
        return result

    days_value = max(1, int(auto_delete_days or 15))
    cutoff_date = (datetime.utcnow() - timedelta(days=days_value)).date()
    try:
        all_files = list_supabase_storage_file_paths("system-logs")
    except Exception:
        result["log_archive_errors"] += 1
        return result

    for object_path in all_files:
        day_key = extract_storage_day_key(object_path)
        if not day_key or day_key >= cutoff_date:
            continue
        deleted = delete_supabase_storage_object_path(object_path)
        if deleted:
            result["deleted_log_archives"] += 1
        else:
            result["log_archive_errors"] += 1
    return result


def collect_supabase_storage_usage(prefix=""):
    total_bytes = 0
    object_count = 0
    folder_count = 0
    queue = deque([(prefix or "").strip("/")])
    visited = set()
    while queue:
        current_prefix = queue.popleft()
        if current_prefix in visited:
            continue
        visited.add(current_prefix)
        for item in list_supabase_storage_objects(current_prefix):
            name = str(item.get("name") or "").strip()
            if not name:
                continue
            metadata = item.get("metadata") or {}
            is_folder = item.get("id") is None and not metadata.get("size")
            if is_folder:
                folder_count += 1
                next_prefix = f"{current_prefix}/{name}".strip("/") if current_prefix else name
                queue.append(next_prefix)
                continue
            size = metadata.get("size")
            try:
                size_int = int(size or 0)
            except Exception:
                size_int = 0
            total_bytes += max(0, size_int)
            object_count += 1
    return {
        "used_bytes": total_bytes,
        "object_count": object_count,
        "folder_count": folder_count,
    }


def build_photo_storage_path(institute, serial, original_filename=""):
    institute_slug = make_storage_slug(institute or "default")
    source_name = os.path.splitext(original_filename or "")[0] or serial or "photo"
    serial_slug = make_storage_slug(source_name)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    return f"{institute_slug}/photos/{serial_slug}-{timestamp}.jpg"


def delete_supabase_storage_url(file_url):
    if not file_url or not is_supabase_enabled():
        return
    base = app.config["SUPABASE_URL"].rstrip("/")
    public_prefix = f"{base}/storage/v1/object/public/"
    if not file_url.startswith(public_prefix):
        return
    remainder = file_url[len(public_prefix):]
    bucket, _, object_path = remainder.partition("/")
    if not bucket or not object_path:
        return
    url = f"{base}/storage/v1/object/{urllib_parse.quote(bucket, safe='')}/{urllib_parse.quote(object_path, safe='/')}"
    req = urllib_request.Request(url, method="DELETE")
    for key, value in supabase_headers().items():
        req.add_header(key, value)
    try:
        with urllib_request.urlopen(req, timeout=30):
            return
    except urllib_error.HTTPError:
        return


def extract_supabase_object_path(file_url):
    if not file_url or not is_supabase_enabled():
        return ""
    base = app.config["SUPABASE_URL"].rstrip("/")
    public_prefix = f"{base}/storage/v1/object/public/"
    if not file_url.startswith(public_prefix):
        return ""
    remainder = file_url[len(public_prefix):]
    bucket, _, object_path = remainder.partition("/")
    if bucket != app.config["SUPABASE_PHOTOS_BUCKET"]:
        return ""
    return object_path


def build_supabase_public_url(object_path, bucket_name=None):
    if not is_supabase_enabled():
        return ""
    bucket = (bucket_name or app.config["SUPABASE_PHOTOS_BUCKET"]).strip() or app.config["SUPABASE_PHOTOS_BUCKET"]
    clean_path = (object_path or "").strip("/")
    if not clean_path:
        return ""
    base = app.config["SUPABASE_URL"].rstrip("/")
    return f"{base}/storage/v1/object/public/{urllib_parse.quote(bucket, safe='')}/{urllib_parse.quote(clean_path, safe='/')}"


def find_last_uploaded_photo_record(institute=None):
    if not is_supabase_enabled():
        return None
    records = list_supabase_records(institute)
    for record in records:
        if record.get("photo_url"):
            return record
    return None


def format_bytes(value):
    units = ["B", "KB", "MB", "GB", "TB"]
    amount = float(max(0, value or 0))
    unit = units[0]
    for unit in units:
        if amount < 1024 or unit == units[-1]:
            break
        amount /= 1024.0
    if unit == "B":
        return f"{int(amount)} {unit}"
    return f"{amount:.2f} {unit}"


def build_supabase_storage_status(institute=None):
    institute = canonicalize_institute_name(institute)
    admin_prefs = load_admin_prefs()
    status = {
        "enabled": is_supabase_enabled(),
        "project_url": app.config["SUPABASE_URL"],
        "bucket_name": app.config["SUPABASE_PHOTOS_BUCKET"],
        "bucket_ready": False,
        "bucket_error": "",
        "institute": institute or "",
        "last_uploaded_file_url": "",
        "last_uploaded_file_path": "",
        "last_uploaded_serial_no": "",
        "storage_quota_mb": admin_prefs.get("storage_quota_mb", 1024),
        "storage_quota_bytes": int(admin_prefs.get("storage_quota_mb", 1024)) * 1024 * 1024,
        "used_bytes": 0,
        "used_display": "0 B",
        "left_bytes": 0,
        "left_display": "0 B",
        "usage_percent": 0,
        "usage_summary": "",
        "object_count": 0,
        "auto_delete_enabled": admin_prefs.get("auto_delete_enabled", False),
        "auto_delete_days": admin_prefs.get("auto_delete_days", 15),
        "last_auto_cleanup_at": admin_prefs.get("last_auto_cleanup_at", ""),
    }
    if not status["enabled"]:
        status["bucket_error"] = "Supabase is not configured"
        return status

    try:
        ensure_supabase_bucket(status["bucket_name"])
        status["bucket_ready"] = True
    except Exception as exc:
        status["bucket_error"] = str(exc) or "Unable to verify bucket"

    if status["bucket_ready"]:
        try:
            usage = collect_supabase_storage_usage("")
            quota_bytes = status["storage_quota_bytes"]
            used_bytes = int(usage.get("used_bytes", 0) or 0)
            left_bytes = max(0, quota_bytes - used_bytes)
            usage_percent = min(100, round((used_bytes / quota_bytes) * 100, 1)) if quota_bytes else 0
            status["used_bytes"] = used_bytes
            status["used_display"] = format_bytes(used_bytes)
            status["left_bytes"] = left_bytes
            status["left_display"] = format_bytes(left_bytes)
            status["usage_percent"] = usage_percent
            status["object_count"] = int(usage.get("object_count", 0) or 0)
            status["usage_summary"] = f"{status['used_display']} used, {status['left_display']} left ({usage_percent}% used)"
        except Exception as exc:
            status["bucket_error"] = status["bucket_error"] or (str(exc) or "Unable to read bucket usage")

    try:
        latest = find_last_uploaded_photo_record(institute or None)
    except Exception as exc:
        status["bucket_error"] = status["bucket_error"] or (str(exc) or "Unable to read latest uploads")
        latest = None

    if latest:
        status["last_uploaded_file_url"] = latest.get("photo_url", "")
        status["last_uploaded_file_path"] = extract_supabase_object_path(latest.get("photo_url", ""))
        status["last_uploaded_serial_no"] = latest.get("serial_no", "")
    return status


def parse_timestamp_display(value):
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def normalize_match_text(value):
    return re.sub(r"\s+", " ", re.sub(r"[_-]+", " ", os.path.splitext(str(value or "").strip().lower())[0])).strip()


def extract_trailing_number_info(value):
    normalized = normalize_match_text(value)
    match = re.match(r"^(.*?)(\d+)\s*$", normalized)
    if not match:
        return {"base": normalized, "number": None}
    return {
        "base": match.group(1).strip(),
        "number": int(match.group(2)),
    }


def sort_records_for_bulk_match(records):
    def sort_key(record):
        info = extract_trailing_number_info(record.get("serial_no", ""))
        number = info["number"] if info["number"] is not None else 10**9
        return (
            number,
            str(record.get("saved_at", "")),
            str(record.get("serial_no", "")),
        )
    return sorted(records, key=sort_key)


def default_batch_name(institute):
    institute_slug = make_storage_slug(institute).replace("_", "-")
    return f"{institute_slug}-{datetime.now().strftime('%Y%m%d-%H%M%S')}"


def pack_supabase_record(record, batch_id, institute, submitted_at, batch_name):
    payload = dict(record)
    payload["institute_name"] = institute
    payload["batch_id"] = batch_id
    payload["batch_name"] = batch_name
    payload["submitted_at"] = submitted_at
    payload["saved_at"] = payload.get("saved_at") or current_timestamp_display()
    return {
        "batch_id": batch_id,
        "institute_name": institute,
        "serial_no": payload.get("serial_no", ""),
        "name": payload.get("name", ""),
        "profile_type": payload.get("profile_type", ""),
        "saved_at": payload.get("saved_at", ""),
        "submitted_at": submitted_at,
        "payload": payload,
    }


def unpack_supabase_record(row):
    payload = dict(row.get("payload") or {})
    payload.setdefault("serial_no", row.get("serial_no", ""))
    payload.setdefault("name", row.get("name", ""))
    payload.setdefault("profile_type", row.get("profile_type", ""))
    payload.setdefault("saved_at", row.get("saved_at", ""))
    payload["submitted_at"] = row.get("submitted_at", payload.get("submitted_at"))
    payload["batch_id"] = row.get("batch_id", payload.get("batch_id"))
    return payload


def list_supabase_batches(institute):
    institute = canonicalize_institute_name(institute)
    if not institute:
        return []
    rows = supabase_request(
        "GET",
        "batches",
        query={
            "select": "id,batch_name,institute_name,status,total_cards,created_at,submitted_at",
            "institute_name": f"eq.{institute}",
            "order": "submitted_at.desc",
        },
    )
    return rows if isinstance(rows, list) else []


def list_all_supabase_batches():
    rows = supabase_request(
        "GET",
        "batches",
        query={
            "select": "id,batch_name,institute_name,status,total_cards,created_at,submitted_at",
            "order": "submitted_at.desc",
        },
    )
    return rows if isinstance(rows, list) else []


def list_supabase_records(institute=None, batch_id=None):
    query = {
        "select": "id,batch_id,institute_name,serial_no,name,profile_type,saved_at,submitted_at,payload",
        "order": "saved_at.asc",
    }
    institute = canonicalize_institute_name(institute)
    if institute:
        query["institute_name"] = f"eq.{institute}"
    if batch_id:
        query["batch_id"] = f"eq.{batch_id}"
    rows = supabase_request("GET", "records", query=query)
    return [unpack_supabase_record(row) for row in (rows if isinstance(rows, list) else [])]


def get_supabase_record_by_serial(institute, serial):
    institute = canonicalize_institute_name(institute)
    serial = (serial or "").strip()
    if not institute or not serial:
        return None
    rows = supabase_request(
        "GET",
        "records",
        query={
            "select": "id,batch_id,institute_name,serial_no,name,profile_type,saved_at,submitted_at,payload",
            "serial_no": f"eq.{serial}",
            "institute_name": f"eq.{institute}",
            "limit": 1,
        },
    )
    if isinstance(rows, list) and rows:
        return unpack_supabase_record(rows[0])
    return None


def find_records_by_serial_lookup(records, lookup):
    lookup_text = str(lookup or "").strip()
    if not lookup_text:
        return []
    exact_matches = [record for record in (records or []) if str(record.get("serial_no", "")).strip() == lookup_text]
    if exact_matches:
        return exact_matches
    return [
        record
        for record in (records or [])
        if str(record.get("serial_no", "")).strip().endswith(lookup_text)
    ]


def normalize_name_lookup(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def find_records_by_name_lookup(records, lookup):
    lookup_text = normalize_name_lookup(lookup)
    if not lookup_text:
        return []
    exact_matches = [
        record for record in (records or [])
        if normalize_name_lookup(record.get("name", "")) == lookup_text
    ]
    if exact_matches:
        return exact_matches
    startswith_matches = [
        record for record in (records or [])
        if normalize_name_lookup(record.get("name", "")).startswith(lookup_text)
    ]
    if startswith_matches:
        return startswith_matches
    return [
        record for record in (records or [])
        if lookup_text in normalize_name_lookup(record.get("name", ""))
    ]


def get_supabase_batch(institute, batch_id):
    institute = canonicalize_institute_name(institute)
    batch_id = (batch_id or "").strip()
    if not institute or not batch_id:
        return None
    rows = supabase_request(
        "GET",
        "batches",
        query={
            "select": "id,batch_name,institute_name,status,total_cards,created_at,submitted_at",
            "id": f"eq.{batch_id}",
            "institute_name": f"eq.{institute}",
            "limit": 1,
        },
    )
    if isinstance(rows, list) and rows:
        return rows[0]
    return None


def build_batch_location_label(records):
    records = records or []
    if not records:
        return ""
    first = records[0] or {}
    facility_location = str(first.get("facility_location", "") or first.get("department", "") or "").strip()
    facility_sub_location = str(first.get("facility_sub_location", "") or "").strip()
    if facility_sub_location and facility_location:
        return f"{facility_sub_location}, {facility_location}"
    return facility_sub_location or facility_location


def enrich_batches_for_overview(batches):
    enriched = []
    for batch in batches or []:
        item = dict(batch or {})
        batch_id = str(item.get("id", "") or "").strip()
        institute = canonicalize_institute_name(item.get("institute_name"))
        location_label = ""
        if batch_id and institute:
            try:
                location_label = build_batch_location_label(list_supabase_records(institute, batch_id))
            except Exception:
                location_label = ""
        item["location_label"] = location_label
        enriched.append(item)
    return enriched


def merge_supabase_batches(institute, batch_ids):
    institute = canonicalize_institute_name(institute)
    unique_batch_ids = []
    for batch_id in batch_ids or []:
        cleaned = str(batch_id or "").strip()
        if cleaned and cleaned not in unique_batch_ids:
            unique_batch_ids.append(cleaned)
    if not institute:
        raise ValueError("Institute is required")
    if len(unique_batch_ids) < 2:
        raise ValueError("Select at least two batches to merge")

    batches = []
    for batch_id in unique_batch_ids:
        batch = get_supabase_batch(institute, batch_id)
        if not batch:
            raise ValueError(f"Batch not found: {batch_id}")
        batches.append(batch)

    target_batch = batches[0]
    target_batch_id = target_batch.get("id", "")
    target_batch_name = target_batch.get("batch_name", "") or target_batch_id
    target_submitted_at = target_batch.get("submitted_at", "") or current_timestamp_display()

    merged_record_total = 0
    merged_batch_names = [target_batch_name]
    for source_batch in batches[1:]:
        source_batch_id = source_batch.get("id", "")
        source_batch_name = source_batch.get("batch_name", "") or source_batch_id
        merged_batch_names.append(source_batch_name)
        source_records = list_supabase_records(institute, source_batch_id)
        merged_record_total += len(source_records)
        for record in source_records:
            serial = str(record.get("serial_no", "")).strip()
            if not serial:
                continue
            payload = dict(record)
            payload["batch_id"] = target_batch_id
            payload["batch_name"] = target_batch_name
            payload["submitted_at"] = target_submitted_at
            supabase_request(
                "PATCH",
                "records",
                payload={
                    "batch_id": target_batch_id,
                    "submitted_at": target_submitted_at,
                    "payload": payload,
                    "name": payload.get("name", ""),
                    "profile_type": payload.get("profile_type", ""),
                    "saved_at": payload.get("saved_at", ""),
                },
                query={
                    "serial_no": f"eq.{serial}",
                    "institute_name": f"eq.{institute}",
                    "batch_id": f"eq.{source_batch_id}",
                },
            )
        supabase_request(
            "DELETE",
            "batches",
            query={"id": f"eq.{source_batch_id}", "institute_name": f"eq.{institute}"},
        )

    target_records = list_supabase_records(institute, target_batch_id)
    supabase_request(
        "PATCH",
        "batches",
        payload={"total_cards": len(target_records)},
        query={"id": f"eq.{target_batch_id}", "institute_name": f"eq.{institute}"},
    )
    return {
        "status": "merged",
        "institute": institute,
        "target_batch_id": target_batch_id,
        "target_batch_name": target_batch_name,
        "merged_batch_ids": unique_batch_ids,
        "merged_batch_names": merged_batch_names,
        "moved_records": merged_record_total,
        "total_records": len(target_records),
    }


def delete_supabase_batch_data(institute, batch_id):
    institute = canonicalize_institute_name(institute)
    batch_id = (batch_id or "").strip()
    if not institute:
        raise ValueError("Institute is required")
    if not batch_id:
        raise ValueError("Batch is required")

    batch = get_supabase_batch(institute, batch_id)
    if not batch:
        raise ValueError("Batch not found")

    batch_records = list_supabase_records(institute, batch_id)
    deleted_photos = 0
    for record in batch_records:
        photo_url = record.get("photo_url", "")
        if photo_url:
            delete_supabase_storage_url(photo_url)
            delete_uploaded_file_from_url(photo_url)
            deleted_photos += 1
        if record.get("photo_drive_id"):
            delete_drive_file(record.get("photo_drive_id"))
        serial_no = record.get("serial_no", "")
        if serial_no:
            delete_local_file_if_exists(os.path.join(UPLOAD_DIR, f"{serial_no}.jpg"))

    supabase_request(
        "DELETE",
        "records",
        query={"batch_id": f"eq.{batch_id}", "institute_name": f"eq.{institute}"},
    )
    supabase_request(
        "DELETE",
        "batches",
        query={"id": f"eq.{batch_id}", "institute_name": f"eq.{institute}"},
    )
    settings = load_settings(institute)
    bindings = sanitize_fabric_batch_background_bindings(settings.get("fabric_batch_background_bindings", {}))
    if batch_id in bindings:
        bindings.pop(batch_id, None)
        settings["fabric_batch_background_bindings"] = bindings
        save_settings(settings, institute)
    return {
        "status": "deleted",
        "batch_id": batch_id,
        "batch_name": batch.get("batch_name", ""),
        "deleted_records": len(batch_records),
        "deleted_photos": deleted_photos,
        "institute": institute,
    }


def save_photo_on_record(target_record, file_storage, institute):
    serial = str(target_record.get("serial_no", "")).strip()
    if not serial:
        raise ValueError("Serial number is required")
    institute = canonicalize_institute_name(institute or target_record.get("institute_name"))
    if not institute:
        raise ValueError("Institute is required")

    old_photo_url = target_record.get("photo_url", "")
    old_drive_id = target_record.get("photo_drive_id")
    photo_url, photo_drive_id = attach_photo_to_serial(file_storage, serial, institute=institute)

    if old_drive_id and old_drive_id != photo_drive_id:
        delete_drive_file(old_drive_id)
    if old_photo_url and old_photo_url != photo_url:
        delete_supabase_storage_url(old_photo_url)
        delete_uploaded_file_from_url(old_photo_url)

    target_record["photo_url"] = photo_url
    target_record["photo_drive_id"] = photo_drive_id or ""
    payload = dict(target_record)
    if is_supabase_enabled():
        query = {"serial_no": f"eq.{serial}", "institute_name": f"eq.{institute}"}
        if payload.get("batch_id"):
            query["batch_id"] = f"eq.{payload.get('batch_id')}"
        supabase_request(
            "PATCH",
            "records",
            payload={
                "payload": payload,
                "saved_at": payload.get("saved_at", ""),
                "name": payload.get("name", ""),
                "profile_type": payload.get("profile_type", ""),
            },
            query=query,
        )
    else:
        records = load_records(institute)
        stored_record = next((rec for rec in records if rec.get("serial_no") == serial), None)
        if not stored_record:
            raise ValueError("Record not found")
        stored_record["photo_url"] = photo_url
        stored_record["photo_drive_id"] = photo_drive_id or ""
        save_records(records, institute)
    return {"serial_no": serial, "photo_url": photo_url, "photo_drive_id": photo_drive_id or ""}


def run_supabase_auto_cleanup(force=False):
    prefs = load_admin_prefs()
    auto_days = max(1, int(prefs.get("auto_delete_days", 15) or 15))
    result = {
        "enabled": prefs.get("auto_delete_enabled", False),
        "auto_delete_days": auto_days,
        "deleted_batches": 0,
        "deleted_records": 0,
        "deleted_photos": 0,
        "deleted_log_archives": 0,
        "log_archive_errors": 0,
        "last_run_at": prefs.get("last_auto_cleanup_at", ""),
    }
    if not is_supabase_enabled() or not prefs.get("auto_delete_enabled"):
        return result

    last_run = parse_timestamp_display(prefs.get("last_auto_cleanup_at"))
    if not force and last_run and (datetime.now() - last_run) < timedelta(hours=1):
        return result

    cutoff = datetime.now() - timedelta(days=auto_days)
    batches = list_all_supabase_batches()
    for batch in batches:
        batch_time = parse_timestamp_display(batch.get("submitted_at") or batch.get("created_at"))
        if not batch_time or batch_time > cutoff:
            continue
        try:
            deleted = delete_supabase_batch_data(batch.get("institute_name"), batch.get("id"))
        except Exception:
            continue
        result["deleted_batches"] += 1
        result["deleted_records"] += int(deleted.get("deleted_records", 0) or 0)
        result["deleted_photos"] += int(deleted.get("deleted_photos", 0) or 0)

    log_cleanup = cleanup_supabase_log_archives(auto_days)
    result["deleted_log_archives"] = int(log_cleanup.get("deleted_log_archives", 0) or 0)
    result["log_archive_errors"] = int(log_cleanup.get("log_archive_errors", 0) or 0)

    result["last_run_at"] = current_timestamp_display()
    prefs["last_auto_cleanup_at"] = result["last_run_at"]
    save_admin_prefs(prefs)
    return result


def create_supabase_batch(institute, records, batch_name=None):
    institute = canonicalize_institute_name(institute)
    if not institute:
        raise ValueError("Institute is required")
    if not records:
        raise ValueError("Batch records are required")
    submitted_at = current_timestamp_display()
    batch_name = (batch_name or "").strip() or default_batch_name(institute)
    batch_rows = supabase_request(
        "POST",
        "batches",
        payload=[{
            "institute_name": institute,
            "batch_name": batch_name,
            "status": "submitted",
            "total_cards": len(records),
            "submitted_at": submitted_at,
            "created_at": submitted_at,
        }],
        prefer_representation=True,
    )
    if not batch_rows or not isinstance(batch_rows, list):
        raise RuntimeError("Unable to create batch")
    batch_id = batch_rows[0].get("id")
    if not batch_id:
        raise RuntimeError("Batch ID missing from Supabase response")
    record_rows = [pack_supabase_record(record, batch_id, institute, submitted_at, batch_name) for record in records]
    supabase_request("POST", "records", payload=record_rows)
    return {
        "batch_id": batch_id,
        "batch_name": batch_name,
        "submitted_at": submitted_at,
        "total_cards": len(records),
        "institute_name": institute,
    }


def normalize_date(value):
    value = (value or "").strip()
    if not value:
        return ""
    if "/" in value:
        parts = value.split("/")
        if len(parts) == 3:
            day, month, year = parts
            return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
        return value
    if "-" in value:
        parts = value.split("-")
        if len(parts) == 3:
            if len(parts[0]) == 4:
                year, month, day = parts
                return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
            day, month, year = parts
            return f"{day.zfill(2)}/{month.zfill(2)}/{year}"
    return value


def current_date_display():
    return datetime.now().strftime("%d/%m/%Y")


def current_timestamp_display():
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def normalize_record_dates(data):
    for field in ("dob", "valid_upto", "inserted_date", "issue_date"):
        if field in data:
            data[field] = normalize_date(data.get(field))
    return data


def canonicalize_institute_name(value):
    value = (value or "").strip()
    if value == "Govt. ANM Training, Jhunjhunu":
        return "Govt. ANM Training Center, Jhunjhunu"
    return value


FACILITY_INSTITUTE_DISPLAY_NAMES = {
    "Govt. Community Health Centre (CHC), Jhunjhunu": "Govt. Community Health Centre (CHC)",
    "Govt. Primary Health Centre (PHC), Jhunjhunu": "Govt. Primary Health Centre (PHC)",
    "Govt. Health Sub Centre, Jhunjhunu": "Govt. Health Sub Centre",
}


def get_display_institute_name(value):
    institute = canonicalize_institute_name(value)
    return FACILITY_INSTITUTE_DISPLAY_NAMES.get(institute, institute)


def format_record_institute_display(record):
    record = record or {}
    institute = get_display_institute_name(record.get("institute_name"))
    facility_location = str(record.get("facility_location", "") or record.get("department", "") or "").strip()
    facility_sub_location = str(record.get("facility_sub_location", "") or "").strip()
    if institute in FACILITY_INSTITUTE_DISPLAY_NAMES.values():
        parts = [institute]
        if facility_sub_location:
            parts.append(facility_sub_location)
        if facility_location:
            parts.append(facility_location)
        return ", ".join(part for part in parts if part)
    return institute


def decorate_record_display(record):
    item = dict(record or {})
    item["display_institute_name"] = format_record_institute_display(item)
    return item


def load_drive_service_account_info():
    raw_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if raw_json:
        return json.loads(raw_json)

    file_path = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE", "").strip()
    if file_path and os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def is_drive_enabled():
    return load_drive_service_account_info() is not None


def get_drive_service():
    global _drive_service
    if _drive_service is not None:
        return _drive_service

    service_account_info = load_drive_service_account_info()
    if not service_account_info:
        return None

    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    credentials = service_account.Credentials.from_service_account_info(
        service_account_info,
        scopes=["https://www.googleapis.com/auth/drive"],
    )
    _drive_service = build("drive", "v3", credentials=credentials, cache_discovery=False)
    return _drive_service


def get_drive_folder_id(kind):
    return (
        app.config.get(f"GOOGLE_DRIVE_{kind.upper()}_FOLDER_ID")
        or app.config["GOOGLE_DRIVE_ROOT_FOLDER_ID"]
        or None
    )


def get_drive_folder_candidates(kind):
    specific = app.config.get(f"GOOGLE_DRIVE_{kind.upper()}_FOLDER_ID") or None
    root = app.config["GOOGLE_DRIVE_ROOT_FOLDER_ID"] or None
    candidates = []
    if specific:
        candidates.append(specific)
    if root and root not in candidates:
        candidates.append(root)
    if not candidates:
        candidates.append(None)
    return candidates


def record_drive_sync_status(filename, synced=None, local_present=None, error=None, file_id=None, folder_id=None):
    if not filename:
        return
    status = _drive_sync_status.get(filename, {})
    status["checked_at"] = current_timestamp_display()
    if synced is not None:
        status["synced"] = bool(synced)
    if local_present is not None:
        status["local_present"] = bool(local_present)
    if error is not None:
        status["error"] = format_drive_error(error)
    if file_id is not None:
        status["file_id"] = file_id or ""
    if folder_id is not None:
        status["folder_id"] = folder_id or ""
    _drive_sync_status[filename] = status


def load_persisted_drive_sync_events():
    if not os.path.exists(DRIVE_SYNC_LOG_FILE):
        return []
    try:
        with open(DRIVE_SYNC_LOG_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return [item for item in data if isinstance(item, dict)]
    except Exception:
        app.logger.warning("Unable to load persisted Drive sync events", exc_info=True)
    return []


def persist_drive_sync_events():
    try:
        with open(DRIVE_SYNC_LOG_FILE, "w", encoding="utf-8") as f:
            json.dump(list(_drive_sync_events), f, indent=2)
    except Exception:
        app.logger.warning("Unable to persist Drive sync events", exc_info=True)


def format_drive_error(error):
    if not error:
        return ""
    if isinstance(error, str):
        return error.strip()
    if error.__class__.__name__ == "HttpError":
        status = ""
        details = ""
        try:
            status = str(getattr(getattr(error, "resp", None), "status", "") or "")
        except Exception:
            status = ""
        try:
            raw_content = getattr(error, "content", b"")
            if isinstance(raw_content, bytes):
                details = raw_content.decode("utf-8", errors="replace").strip()
            else:
                details = str(raw_content or "").strip()
        except Exception:
            details = ""
        parts = ["HttpError"]
        if status:
            parts.append(f"status={status}")
        if details:
            parts.append(details)
        return " | ".join(parts)
    return f"{error.__class__.__name__}: {str(error).strip()}"


def execute_drive_request(request, retries=3):
    return request.execute(num_retries=retries)


def append_drive_sync_event(filename, institute="", kind="", status="", error="", file_id="", folder_id=""):
    event = {
        "time": current_timestamp_display(),
        "filename": filename or "",
        "institute": canonicalize_institute_name(institute) or "",
        "kind": kind or "",
        "status": status or "",
        "error": format_drive_error(error),
        "file_id": file_id or "",
        "folder_id": folder_id or "",
    }
    _drive_sync_events.appendleft(event)
    persist_drive_sync_events()
    message = f"Drive sync event status={event['status']} file={event['filename'] or '-'} institute={event['institute'] or '-'}"
    if event["error"]:
        app.logger.warning("%s error=%s", message, event["error"])
    else:
        app.logger.info("%s", message)


def extract_institute_from_filename(filename):
    name = (filename or "").strip()
    if "__" not in name:
        return ""
    try:
        slug = name.split("__", 1)[1].rsplit(".json", 1)[0]
    except Exception:
        return ""
    return slug.replace("_", " ").strip()


def get_recent_drive_sync_events(institute=None):
    institute = canonicalize_institute_name(institute)
    if not institute:
        return list(_drive_sync_events)
    slug = make_storage_slug(institute)
    prefixes = (
        f"records__{slug}",
        f"certificates__{slug}",
        f"settings__{slug}",
    )
    return [event for event in _drive_sync_events if (event.get("filename") or "").startswith(prefixes)]


def make_drive_public(service, file_id):
    execute_drive_request(service.permissions().create(
        fileId=file_id,
        body={"type": "anyone", "role": "reader"},
        fields="id",
    ))


def build_drive_view_url(file_id):
    return f"https://drive.google.com/uc?export=view&id={file_id}"


def upload_bytes_to_drive(file_bytes, filename, mime_type, kind):
    service = get_drive_service()
    if service is None:
        return None, None

    from googleapiclient.http import MediaIoBaseUpload

    metadata = {"name": filename}
    folder_id = get_drive_folder_id(kind)
    if folder_id:
        metadata["parents"] = [folder_id]

    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type, resumable=False)
    created = execute_drive_request(service.files().create(body=metadata, media_body=media, fields="id"))
    file_id = created["id"]
    make_drive_public(service, file_id)
    return file_id, build_drive_view_url(file_id)


def _drive_query_string(value):
    return value.replace("\\", "\\\\").replace("'", "\\'")


def find_drive_file_id(filename, kind):
    service = get_drive_service()
    if service is None:
        return None

    for folder_id in get_drive_folder_candidates(kind):
        query = [f"name='{_drive_query_string(filename)}'", "trashed=false"]
        if folder_id:
            query.append(f"'{_drive_query_string(folder_id)}' in parents")
        try:
            response = execute_drive_request(service.files().list(
                q=" and ".join(query),
                fields="files(id, name, parents)",
                pageSize=1,
            ))
        except Exception:
            continue
        files = response.get("files", [])
        if files:
            found = files[0]
            resolved_folder_id = (found.get("parents") or [folder_id or ""])[0] if found.get("parents") or folder_id else ""
            record_drive_sync_status(filename, synced=True, error="", file_id=found.get("id", ""), folder_id=resolved_folder_id)
            return found["id"]
    return None


def download_drive_file_bytes(file_id):
    service = get_drive_service()
    if service is None or not file_id:
        return None

    from googleapiclient.http import MediaIoBaseDownload

    buffer = io.BytesIO()
    request = service.files().get_media(fileId=file_id)
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk(num_retries=3)
    return buffer.getvalue()


def upsert_private_drive_file(file_bytes, filename, mime_type, kind):
    service = get_drive_service()
    if service is None:
        return None

    from googleapiclient.http import MediaIoBaseUpload

    file_id = _drive_json_file_ids.get(filename) or find_drive_file_id(filename, kind)
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type, resumable=False)
    if file_id:
        execute_drive_request(service.files().update(fileId=file_id, media_body=media))
        _drive_json_file_ids[filename] = file_id
        record_drive_sync_status(filename, synced=True, error="", file_id=file_id)
        append_drive_sync_event(
            filename,
            institute=extract_institute_from_filename(filename),
            kind=kind,
            status="drive_synced",
            file_id=file_id,
        )
        return file_id

    last_error = None
    for folder_id in get_drive_folder_candidates(kind):
        metadata = {"name": filename}
        if folder_id:
            metadata["parents"] = [folder_id]
        media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mime_type, resumable=False)
        try:
            created = execute_drive_request(service.files().create(body=metadata, media_body=media, fields="id, parents"))
            file_id = created["id"]
            _drive_json_file_ids[filename] = file_id
            resolved_folder_id = (created.get("parents") or [folder_id or ""])[0] if created.get("parents") or folder_id else ""
            record_drive_sync_status(filename, synced=True, error="", file_id=file_id, folder_id=resolved_folder_id)
            append_drive_sync_event(
                filename,
                institute=extract_institute_from_filename(filename),
                kind=kind,
                status="drive_synced",
                file_id=file_id,
                folder_id=resolved_folder_id,
            )
            return file_id
        except Exception as exc:
            last_error = exc
            continue
    if last_error:
        record_drive_sync_status(filename, synced=False, error=last_error)
        append_drive_sync_event(
            filename,
            institute=extract_institute_from_filename(filename),
            kind=kind,
            status="drive_failed",
            error=last_error,
        )
        raise last_error
    return None


def load_json_store(local_path, drive_filename, default_value):
    if is_drive_enabled():
        try:
            file_id = _drive_json_file_ids.get(drive_filename) or find_drive_file_id(drive_filename, "data")
            if file_id:
                _drive_json_file_ids[drive_filename] = file_id
                file_bytes = download_drive_file_bytes(file_id)
                if file_bytes is not None:
                    with open(local_path, "wb") as f:
                        f.write(file_bytes)
                    record_drive_sync_status(
                        drive_filename,
                        synced=True,
                        local_present=True,
                        error="",
                        file_id=file_id,
                    )
                    append_drive_sync_event(
                        drive_filename,
                        institute=extract_institute_from_filename(drive_filename),
                        kind="data",
                        status="drive_synced",
                        file_id=file_id,
                    )
                    return json.loads(file_bytes.decode("utf-8"))
            elif os.path.exists(local_path):
                with open(local_path, "rb") as f:
                    existing_bytes = f.read()
                if existing_bytes:
                    synced_file_id = upsert_private_drive_file(existing_bytes, drive_filename, "application/json", "data")
                    record_drive_sync_status(
                        drive_filename,
                        synced=True,
                        local_present=True,
                        error="",
                        file_id=synced_file_id,
                    )
                    append_drive_sync_event(
                        drive_filename,
                        institute=extract_institute_from_filename(drive_filename),
                        kind="data",
                        status="drive_synced",
                        file_id=synced_file_id,
                    )
                    return json.loads(existing_bytes.decode("utf-8"))
        except Exception:
            record_drive_sync_status(
                drive_filename,
                synced=False,
                local_present=os.path.exists(local_path),
                error="Drive sync failed while loading this file.",
            )
            append_drive_sync_event(
                drive_filename,
                institute=extract_institute_from_filename(drive_filename),
                kind="data",
                status="local_only",
                error="Drive sync failed while loading this file.",
            )
            app.logger.warning("Drive sync failed while loading %s", drive_filename, exc_info=True)

    if os.path.exists(local_path):
        with open(local_path, "r", encoding="utf-8") as f:
            try:
                record_drive_sync_status(
                    drive_filename,
                    synced=not is_drive_enabled(),
                    local_present=True,
                    error="" if not is_drive_enabled() else _drive_sync_status.get(drive_filename, {}).get("error", ""),
                )
                if is_drive_enabled() and _drive_sync_status.get(drive_filename, {}).get("error", ""):
                    append_drive_sync_event(
                        drive_filename,
                        institute=extract_institute_from_filename(drive_filename),
                        kind="data",
                        status="local_only",
                        error=_drive_sync_status.get(drive_filename, {}).get("error", ""),
                    )
                return json.load(f)
            except json.JSONDecodeError:
                pass

    return json.loads(json.dumps(default_value))


def save_json_store(local_path, drive_filename, payload):
    json_text = json.dumps(payload, indent=2)
    with open(local_path, "w", encoding="utf-8") as f:
        f.write(json_text)

    if is_drive_enabled():
        try:
            file_id = upsert_private_drive_file(json_text.encode("utf-8"), drive_filename, "application/json", "data")
            record_drive_sync_status(drive_filename, synced=True, local_present=True, error="", file_id=file_id)
        except Exception as exc:
            record_drive_sync_status(drive_filename, synced=False, local_present=True, error=exc)
            append_drive_sync_event(
                drive_filename,
                institute=extract_institute_from_filename(drive_filename),
                kind="data",
                status="local_only",
                error=exc,
            )
            app.logger.warning("Drive sync failed while saving %s", drive_filename, exc_info=True)
    else:
        record_drive_sync_status(drive_filename, synced=False, local_present=True, error="")
        append_drive_sync_event(
            drive_filename,
            institute=extract_institute_from_filename(drive_filename),
            kind="data",
            status="local_only",
            error="Google Drive is not configured.",
        )


def build_drive_storage_status(institute=None):
    institute = canonicalize_institute_name(institute)
    data_folder_id = get_drive_folder_id("data")
    files = []
    targets = (
        ("records", make_storage_filename("records", institute)) if institute else ("records", ""),
        ("certificates", make_storage_filename("certificates", institute)) if institute else ("certificates", ""),
        ("settings", make_storage_filename("settings", institute)) if institute else ("settings", ""),
    )
    for kind, filename in targets:
        file_id = None
        local_path = make_storage_path(kind, institute) if institute and filename else ""
        local_present = bool(local_path and os.path.exists(local_path))
        last_error = _drive_sync_status.get(filename, {}).get("error", "")
        resolved_folder_id = _drive_sync_status.get(filename, {}).get("folder_id", "")
        try:
            file_id = _drive_json_file_ids.get(filename) or (find_drive_file_id(filename, "data") if filename else None)
            if file_id:
                _drive_json_file_ids[filename] = file_id
                resolved_folder_id = _drive_sync_status.get(filename, {}).get("folder_id", resolved_folder_id)
                last_error = _drive_sync_status.get(filename, {}).get("error", "")
        except Exception as exc:
            last_error = str(exc)
            app.logger.warning("Unable to resolve Drive file status for %s", filename, exc_info=True)
        status = "drive_synced" if file_id else ("local_only" if local_present else ("select_institute" if not institute else "missing"))
        files.append({
            "kind": kind,
            "name": filename or f"{kind} file requires institute selection",
            "present": bool(file_id),
            "file_id": file_id or "",
            "local_present": local_present,
            "status": status,
            "last_error": last_error,
            "folder_id": resolved_folder_id,
        })

    return {
        "drive_enabled": is_drive_enabled(),
        "data_folder_id": data_folder_id or "",
        "institute": institute or "",
        "files": files,
        "events": get_recent_drive_sync_events(institute),
    }


# Local-only storage mode: keep all JSON and assets on the app filesystem.
def is_drive_enabled():
    return False


def get_drive_service():
    return None


def list_drive_storage_filenames(kind):
    return []


def load_json_store(local_path, drive_filename, default_value):
    if os.path.exists(local_path):
        with open(local_path, "r", encoding="utf-8") as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                pass
    return json.loads(json.dumps(default_value))


def save_json_store(local_path, drive_filename, payload):
    with open(local_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)


def build_drive_storage_status(institute=None):
    institute = canonicalize_institute_name(institute)
    return {
        "drive_enabled": False,
        "data_folder_id": "",
        "institute": institute or "",
        "files": [],
        "events": [],
    }


def delete_drive_file(file_id):
    if not file_id:
        return
    service = get_drive_service()
    if service is None:
        return
    try:
        execute_drive_request(service.files().delete(fileId=file_id))
    except Exception:
        pass


def delete_local_file_if_exists(path):
    if path and os.path.exists(path):
        os.remove(path)


def delete_uploaded_file_from_url(file_url):
    if not file_url:
        return
    delete_supabase_storage_url(file_url)
    marker = "/uploads/"
    if marker not in file_url:
        return
    filename = file_url.split(marker, 1)[1].split("?", 1)[0].strip("/")
    if filename:
        delete_local_file_if_exists(os.path.join(UPLOAD_DIR, filename))


def delete_generated_asset_from_url(file_url):
    if not file_url:
        return
    delete_supabase_storage_url(file_url)
    marker = "/generated-assets/"
    if marker not in file_url:
        return
    filename = file_url.split(marker, 1)[1].split("?", 1)[0].strip("/")
    if filename:
        delete_local_file_if_exists(os.path.join(ASSET_DIR, filename))


def download_drive_file(file_id):
    if not file_id:
        return None
    service = get_drive_service()
    if service is None:
        return None

    from googleapiclient.http import MediaIoBaseDownload

    request = service.files().get_media(fileId=file_id)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk(num_retries=3)
    return buffer.getvalue()


def download_file_from_url(file_url):
    if not file_url:
        return None
    try:
        with urllib_request.urlopen(file_url, timeout=60) as response:
            return response.read()
    except Exception:
        return None


def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("admin_authenticated"):
            return jsonify({"error": "Unauthorized"}), 401
        return view_func(*args, **kwargs)
    return wrapped


def verify_cron_request():
    configured = (app.config.get("CRON_BACKUP_TOKEN") or "").strip()
    auth_header = (request.headers.get("authorization") or "").strip()
    bearer_value = ""
    if auth_header.lower().startswith("bearer "):
        bearer_value = auth_header[7:].strip()
    # If no token is configured, only allow Vercel Cron-originated calls.
    if not configured:
        return request.headers.get("x-vercel-cron") == "1"
    provided = (request.headers.get("x-cron-token") or request.args.get("token") or bearer_value).strip()
    return provided == configured


def make_asset_slug(value):
    safe = secure_filename((value or "").strip())
    return safe or "default"


def get_filtered_records():
    institute = canonicalize_institute_name(request.args.get("institute"))
    batch_id = (request.args.get("batch_id") or "").strip()
    if is_supabase_enabled():
        records = list_supabase_records(institute, batch_id or None)
    else:
        records = load_records(institute)
    records = [decorate_record_display(record) for record in records]
    return records, institute


def summarize_batch(records):
    saved = [r for r in records if not r.get("submitted_at")]
    submitted = [r for r in records if r.get("submitted_at")]
    return {
        "saved_count": len(saved),
        "submitted_count": len(submitted),
        "total_count": len(records),
    }

def make_institute_code(value):
    institute = canonicalize_institute_name(value)
    if not institute:
        return "GEN"
    words = [word for word in secure_filename(institute).replace("_", " ").split() if word]
    initials = "".join(word[0] for word in words[:6]).upper()
    if len(initials) >= 3:
        return initials[:6]
    compact = "".join(ch for ch in secure_filename(institute).upper() if ch.isalnum())
    return (compact[:6] or "GEN")


def gen_serial(institute=None):
    date_part = datetime.now().strftime("%Y%m%d")
    rand_part = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    institute_code = make_institute_code(institute)
    return f"ID-{institute_code}-{date_part}-{rand_part}"


def gen_certificate_no():
    date_part = datetime.now().strftime("%Y%m%d")
    rand_part = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
    return f"CERT-{date_part}-{rand_part}"

def detect_primary_face(img):
    try:
        import cv2
        import numpy as np
    except Exception:
        return None

    rgb = np.array(img)
    gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)
    cascade_path = os.path.join(cv2.data.haarcascades, "haarcascade_frontalface_default.xml")
    face_cascade = cv2.CascadeClassifier(cascade_path)
    faces = face_cascade.detectMultiScale(
        gray,
        scaleFactor=1.1,
        minNeighbors=5,
        minSize=(60, 60),
    )
    if len(faces) == 0:
        return None
    return max(faces, key=lambda face: face[2] * face[3])


def crop_passport_frame(img, face_box=None):
    w, h = img.size
    target_ratio = 3 / 4

    if face_box is not None:
        fx, fy, fw, fh = [int(v) for v in face_box]
        crop_h = min(h, max(int(fh * 3.2), 220))
        crop_w = int(crop_h * target_ratio)

        if crop_w > w:
            crop_w = w
            crop_h = int(crop_w / target_ratio)

        center_x = fx + fw // 2
        face_center_y = fy + fh // 2
        center_y = int(face_center_y + fh * 0.55)

        left = max(0, min(center_x - crop_w // 2, w - crop_w))
        top = max(0, min(center_y - crop_h // 2, h - crop_h))
        return img.crop((left, top, left + crop_w, top + crop_h))

    current_ratio = w / h
    if current_ratio > target_ratio:
        new_w = int(h * target_ratio)
        left = (w - new_w) // 2
        return img.crop((left, 0, left + new_w, h))

    new_h = int(w / target_ratio)
    top = max(0, int((h - new_h) * 0.16))
    return img.crop((0, top, w, min(h, top + new_h)))


def autocrop_passport(img_path, out_path):
    """Create a 3:4 passport crop while preserving high resolution."""
    img = Image.open(img_path)
    img = ImageOps.exif_transpose(img).convert("RGB")
    face_box = detect_primary_face(img)
    img = crop_passport_frame(img, face_box)
    img = normalize_passport_resolution(img)
    img.save(out_path, "JPEG", quality=95, optimize=True, progressive=True)


def normalize_passport_resolution(img):
    """
    Preserve details for print use:
    - keep original resolution in normal cases
    - reduce only very large uploads
    - gently upscale very tiny crops
    """
    width, height = img.size
    max_height = 1600
    min_height = 720

    if height > max_height:
        scale = max_height / float(height)
        new_size = (max(1, int(width * scale)), max_height)
        return img.resize(new_size, Image.LANCZOS)

    if height < min_height:
        scale = min_height / float(height)
        new_size = (max(1, int(width * scale)), min_height)
        return img.resize(new_size, Image.LANCZOS)

    return img

def save_image_asset(file_storage, filename, kind):
    try:
        file_storage.stream.seek(0)
    except Exception:
        pass

    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("Empty image upload")

    try:
        img = Image.open(io.BytesIO(file_bytes))
        img = ImageOps.exif_transpose(img).convert("RGB")
    except UnidentifiedImageError as exc:
        raise ValueError("Unsupported image format") from exc

    local_path = os.path.join(ASSET_DIR, filename)
    output = io.BytesIO()
    img.save(output, "JPEG", quality=92)
    output_bytes = output.getvalue()
    with open(local_path, "wb") as image_file:
        image_file.write(output_bytes)

    drive_id = None
    drive_url = None
    if is_drive_enabled():
        try:
            drive_id, drive_url = upload_bytes_to_drive(output_bytes, filename, "image/jpeg", kind)
        except Exception:
            app.logger.warning("Drive upload failed for asset %s", filename, exc_info=True)
    return drive_url or f"/generated-assets/{filename}", drive_id


def save_background_image(file_storage, institute_name):
    filename = f"card_background_{make_asset_slug(institute_name)}.jpg"
    return save_image_asset(file_storage, filename, "backgrounds")


def save_print_background_image(file_storage, institute_name, background_id):
    filename = build_print_background_filename(institute_name, background_id)
    return save_image_asset(file_storage, filename, "backgrounds")


def save_certificate_background_image(file_storage, institute_name):
    filename = f"certificate_background_{make_asset_slug(institute_name)}.jpg"
    return save_image_asset(file_storage, filename, "backgrounds")


def save_signature_image(file_storage, institute_name):
    img = Image.open(file_storage.stream)
    img = ImageOps.exif_transpose(img)
    if img.mode not in ("RGBA", "LA"):
        img = img.convert("RGBA")
    img = img.convert("RGBA")
    pixels = []
    for r, g, b, a in img.getdata():
        if a == 0:
            pixels.append((r, g, b, a))
            continue
        brightness = (r + g + b) / 3
        channel_spread = max(r, g, b) - min(r, g, b)
        if brightness > 235 and channel_spread < 22:
            pixels.append((255, 255, 255, 0))
            continue
        ink_boost = max(0, int((235 - brightness) * 1.35))
        alpha = max(70, min(255, ink_boost + 70))
        pixels.append((20, 20, 20, alpha))
    img.putdata(pixels)
    bbox = img.getbbox()
    if bbox:
        img = img.crop(bbox)
    filename = f"hod_signature_{make_asset_slug(institute_name)}.png"
    local_path = os.path.join(ASSET_DIR, filename)
    img.save(local_path, "PNG")
    drive_id = None
    drive_url = None
    if is_drive_enabled():
        with open(local_path, "rb") as f:
            drive_id, drive_url = upload_bytes_to_drive(f.read(), filename, "image/png", "signatures")
    return drive_url or f"/generated-assets/{filename}", drive_id


def save_print_studio_asset(file_storage, institute_name, asset_type):
    valid_types = {"background", "photo", "logo", "signature"}
    kind = str(asset_type or "").strip().lower()
    if kind not in valid_types:
        raise ValueError("Invalid asset type")

    try:
        file_storage.stream.seek(0)
    except Exception:
        pass
    file_bytes = file_storage.read()
    if not file_bytes:
        raise ValueError("Empty image upload")

    try:
        img = Image.open(io.BytesIO(file_bytes))
        img = ImageOps.exif_transpose(img)
    except UnidentifiedImageError as exc:
        raise ValueError("Unsupported image format") from exc

    safe_institute = make_asset_slug(institute_name)
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
    base_name = f"print_studio_{kind}_{safe_institute}_{timestamp}"

    if kind == "signature":
        # Make signature print-friendly:
        # 1) transparent background (remove near-white paper)
        # 2) darker ink with preserved anti-aliased edges
        # 3) slight sharpness boost
        rgba = img.convert("RGBA")
        processed = []
        for r, g, b, a in rgba.getdata():
            if a == 0:
                processed.append((r, g, b, 0))
                continue
            brightness = (r + g + b) / 3
            channel_spread = max(r, g, b) - min(r, g, b)
            if brightness > 236 and channel_spread < 20:
                processed.append((255, 255, 255, 0))
                continue
            ink_boost = max(0, int((235 - brightness) * 1.35))
            alpha = max(72, min(255, ink_boost + 72))
            processed.append((16, 16, 16, alpha))
        rgba.putdata(processed)
        bbox = rgba.getbbox()
        if bbox:
            rgba = rgba.crop(bbox)
        rgba = ImageEnhance.Sharpness(rgba).enhance(1.18)
        mime_type = "image/png"
        filename = f"{base_name}.png"
        output = io.BytesIO()
        rgba.save(output, "PNG")
    else:
        mime_type = "image/jpeg"
        filename = f"{base_name}.jpg"
        output = io.BytesIO()
        img.convert("RGB").save(output, "JPEG", quality=94)

    output_bytes = output.getvalue()
    local_path = os.path.join(ASSET_DIR, filename)
    with open(local_path, "wb") as image_file:
        image_file.write(output_bytes)

    if is_supabase_enabled():
        object_path = f"{make_storage_slug(institute_name or 'default')}/print-studio/{kind}/{filename}"
        return upload_bytes_to_supabase_storage(output_bytes, object_path, mime_type), ""

    drive_id = None
    drive_url = None
    if is_drive_enabled():
        try:
            drive_id, drive_url = upload_bytes_to_drive(output_bytes, filename, mime_type, "assets")
        except Exception:
            app.logger.warning("Drive upload failed for print studio asset %s", filename, exc_info=True)
    return drive_url or f"/generated-assets/{filename}", drive_id


def build_placeholder_avatar_bytes(record=None, sequence_no=None):
    width, height = 300, 400
    image = Image.new("RGB", (width, height), "#f3efe7")
    draw = ImageDraw.Draw(image)

    draw.rounded_rectangle((18, 18, width - 18, height - 18), radius=24, fill="#fbf8f2", outline="#d8cfc1", width=3)
    draw.ellipse((92, 62, 208, 178), fill="#cfb08c")
    draw.ellipse((74, 172, 226, 340), fill="#4c6a92")
    draw.rectangle((74, 246, 226, 340), fill="#4c6a92")
    draw.ellipse((126, 162, 174, 214), fill="#cfb08c")

    approx_char_width = 7
    if sequence_no is not None:
        seq_text = f"S.No {sequence_no}"
        seq_x = max(24, (width - len(seq_text) * approx_char_width) // 2)
        draw.text((seq_x, 326), seq_text, fill="#8b0000")

    label = "PHOTO NOT UPLOADED"
    label_x = max(24, (width - len(label) * approx_char_width) // 2)
    draw.text((label_x, 352), label, fill="#7a4d24")

    if record and str(record.get("name", "")).strip():
        name_text = str(record.get("name", "")).strip()[:24]
        name_x = max(24, (width - len(name_text) * approx_char_width) // 2)
        draw.text((name_x, 28), name_text, fill="#2c1810")

    output = io.BytesIO()
    image.save(output, "JPEG", quality=92)
    return output.getvalue()


def build_export_csv_text(records):
    csv_buf = io.StringIO()
    writer = csv.writer(csv_buf)
    writer.writerow([
        "S.No", "Serial No", "Profile Type", "Name", "Course/Designation", "Batch", "Father Name", "Aadhaar No.",
        "Employee ID", "Department", "Date of Birth", "Contact No", "Blood Group", "Address", "Valid Upto",
        "Institute", "Photo File", "Photo URL", "Saved At", "Submitted At"
    ])
    for index, rec in enumerate(records, 1):
        writer.writerow([
            index,
            rec.get("serial_no", ""),
            rec.get("profile_type", ""),
            rec.get("name", ""),
            rec.get("training_year") or rec.get("course") or rec.get("designation", ""),
            rec.get("batch_session", ""),
            rec.get("father_name", ""),
            rec.get("aadhaar_no", ""),
            rec.get("employee_id", ""),
            rec.get("department", ""),
            rec.get("dob", ""),
            rec.get("contact", ""),
            rec.get("blood_group", ""),
            rec.get("address", ""),
            rec.get("valid_upto", ""),
            rec.get("institute_name", ""),
            f"{index}.jpg",
            rec.get("photo_url", ""),
            rec.get("saved_at", ""),
            rec.get("submitted_at", ""),
        ])
    return csv_buf.getvalue()


def build_export_excel_bytes(records):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "ID Card Records"
    headers = ["S.No", "Serial No", "Profile Type", "Name", "Course/Designation", "Batch", "Father Name", "Aadhaar No.",
               "Employee ID", "Department", "Date of Birth", "Contact No", "Blood Group", "Address", "Valid Upto",
               "Institute", "Photo File", "Saved At"]
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row, rec in enumerate(records, 2):
        sequence_no = row - 1
        ws.cell(row=row, column=1, value=sequence_no)
        ws.cell(row=row, column=2, value=rec.get("serial_no", ""))
        ws.cell(row=row, column=3, value=rec.get("profile_type", ""))
        ws.cell(row=row, column=4, value=rec.get("name", ""))
        ws.cell(row=row, column=5, value=rec.get("training_year") or rec.get("course") or rec.get("designation", ""))
        ws.cell(row=row, column=6, value=rec.get("batch_session", ""))
        ws.cell(row=row, column=7, value=rec.get("father_name", ""))
        ws.cell(row=row, column=8, value=rec.get("aadhaar_no", ""))
        ws.cell(row=row, column=9, value=rec.get("employee_id", ""))
        ws.cell(row=row, column=10, value=rec.get("department", ""))
        ws.cell(row=row, column=11, value=rec.get("dob", ""))
        ws.cell(row=row, column=12, value=rec.get("contact", ""))
        ws.cell(row=row, column=13, value=rec.get("blood_group", ""))
        ws.cell(row=row, column=14, value=rec.get("address", ""))
        ws.cell(row=row, column=15, value=rec.get("valid_upto", ""))
        ws.cell(row=row, column=16, value=rec.get("institute_name", ""))
        ws.cell(row=row, column=17, value=f"{sequence_no}.jpg")
        ws.cell(row=row, column=18, value=rec.get("saved_at", ""))
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_export_xls_bytes(records):
    headers = ["S.No", "Serial No", "Profile Type", "Name", "Course/Designation", "Batch", "Father Name", "Aadhaar No.",
               "Employee ID", "Department", "Date of Birth", "Contact No", "Blood Group", "Address", "Valid Upto",
               "Institute", "Photo File", "Saved At"]
    rows = []
    for index, rec in enumerate(records, 1):
        rows.append([
            index,
            rec.get("serial_no", ""),
            rec.get("profile_type", ""),
            rec.get("name", ""),
            rec.get("training_year") or rec.get("course") or rec.get("designation", ""),
            rec.get("batch_session", ""),
            rec.get("father_name", ""),
            rec.get("aadhaar_no", ""),
            rec.get("employee_id", ""),
            rec.get("department", ""),
            rec.get("dob", ""),
            rec.get("contact", ""),
            rec.get("blood_group", ""),
            rec.get("address", ""),
            rec.get("valid_upto", ""),
            rec.get("institute_name", ""),
            f"{index}.jpg",
            rec.get("saved_at", ""),
        ])

    def html_cell(tag, value):
        text_value = escape(str(value or ""))
        if tag == "th":
            return f"<{tag} style=\"mso-number-format:'\\@';\">{text_value}</{tag}>"
        return f"<{tag} style=\"mso-number-format:'\\@'; white-space:nowrap;\">{text_value}</{tag}>"

    html = [
        "<html>",
        "<head>",
        '<meta http-equiv="Content-Type" content="text/html; charset=utf-8">',
        "</head>",
        "<body>",
        "<table border='1' style=\"border-collapse:collapse;\">",
        "<tr>",
    ]
    html.extend(html_cell("th", header) for header in headers)
    html.append("</tr>")
    for row in rows:
        html.append("<tr>")
        html.extend(html_cell("td", value) for value in row)
        html.append("</tr>")
    html.extend(["</table>", "</body>", "</html>"])
    return "".join(html).encode("utf-8")

@app.route("/")
def home():
    return render_template("home.html")


@app.route("/id-card")
def index():
    return render_template("index.html")


@app.route("/certificate")
def certificate():
    return render_template("certificate.html")

@app.route("/admin")
def admin():
    if not session.get("admin_authenticated"):
        return render_template("admin_login.html")
    return render_template("admin.html")


@app.route("/admin/print-cards")
@admin_required
def admin_print_cards():
    records, institute = get_filtered_records()
    batch_id = (request.args.get("batch_id") or "").strip()
    settings = load_settings(institute)
    auto_print = (request.args.get("autoprint") or "").strip() == "1"
    return render_template(
        "print_cards.html",
        records=records,
        institute=institute,
        display_institute_name=get_display_institute_name(institute),
        batch_id=batch_id,
        settings=settings,
        auto_print=auto_print,
    )


def build_print_preview_context():
    records, institute = get_filtered_records()
    batch_id = (request.args.get("batch_id") or "").strip()
    settings = load_settings(institute)
    sample_record = records[0] if records else {
        "profile_type": "student",
        "name": "Aarav Sharma",
        "training_year": "B.Sc. Nursing - Year 2",
        "blood_group": "B+",
        "batch_session": "2025-26",
        "father_name": "Mahesh Sharma",
        "dob": "14/08/2004",
        "contact": "9876543210",
        "valid_upto": "31/03/2027",
        "serial_no": "ID-20260408-DEMO",
        "institute_name": institute or "Govt. Medical College, Jhunjhunu",
        "department": "Nursing",
        "employee_id": "EMP-102",
    }
    return {
        "records": records,
        "institute": institute,
        "batch_id": batch_id,
        "settings": settings,
        "sample_record": sample_record,
        "build_tag": app.config.get("APP_BUILD_TAG", ""),
    }


@app.route("/admin/print-studio")
@admin_required
def admin_print_studio():
    response = make_response(render_template("print_studio_fabric.html", **build_print_preview_context()))
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["X-App-Build"] = app.config.get("APP_BUILD_TAG", "")
    return response


@app.route("/admin/backgrounds")
@admin_required
def admin_backgrounds():
    response = make_response(render_template("background_manager.html"))
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    response.headers["X-App-Build"] = app.config.get("APP_BUILD_TAG", "")
    return response


@app.route("/api/fabric-design", methods=["GET", "POST"])
@admin_required
def fabric_design():
    json_payload = request.get_json(silent=True) if request.is_json else {}
    institute = canonicalize_institute_name(
        request.args.get("institute")
        or (json_payload.get("institute") if isinstance(json_payload, dict) else "")
        or request.form.get("institute")
    )
    settings = load_settings(institute)
    if request.method == "GET":
        return jsonify({
            "institute": institute,
            "design": settings.get("fabric_design", {}),
        })

    payload = json_payload if isinstance(json_payload, dict) else {}
    if not institute:
        return jsonify({"error": "Institute is required to save design"}), 400
    design = payload.get("design")
    if not isinstance(design, dict):
        return jsonify({"error": "Invalid design payload"}), 400
    serialized = json.dumps(design, ensure_ascii=False)
    if len(serialized) > 2_500_000:
        return jsonify({"error": "Design payload too large"}), 400
    settings["fabric_design"] = design
    save_settings(settings, institute)
    return jsonify({"saved": True, "institute": institute})


@app.route("/api/fabric-assets/upload", methods=["POST"])
@admin_required
def fabric_asset_upload():
    institute = canonicalize_institute_name(request.form.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    if "asset" not in request.files:
        return jsonify({"error": "No file"}), 400
    file_storage = request.files["asset"]
    if not secure_filename(file_storage.filename):
        return jsonify({"error": "Invalid filename"}), 400
    asset_type = (request.form.get("asset_type") or "").strip().lower()
    try:
        asset_url, _ = save_print_studio_asset(file_storage, institute, asset_type)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception:
        return jsonify({"error": "Unable to upload print studio asset"}), 500
    return jsonify({
        "status": "saved",
        "institute": institute,
        "asset_type": asset_type,
        "url": asset_url,
    })


@app.route("/api/fabric-assets/delete", methods=["POST"])
@admin_required
def fabric_asset_delete():
    payload = request.get_json(silent=True) or {}
    institute = canonicalize_institute_name(payload.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    file_url = str(payload.get("url") or "").strip()
    if not file_url:
        return jsonify({"error": "Asset URL is required"}), 400

    deleted = False
    if "/generated-assets/" in file_url:
        filename = file_url.split("/generated-assets/", 1)[1].split("?", 1)[0].strip("/")
        if not filename.startswith("print_studio_"):
            return jsonify({"error": "Only print studio assets can be deleted"}), 400
        delete_generated_asset_from_url(file_url)
        deleted = True
    else:
        object_path = extract_supabase_object_path(file_url)
        if object_path:
            if "/print-studio/" not in object_path:
                return jsonify({"error": "Only print studio assets can be deleted"}), 400
            delete_supabase_storage_url(file_url)
            deleted = True

    return jsonify({
        "status": "deleted" if deleted else "ignored",
        "deleted": deleted,
        "url": file_url,
        "institute": institute,
    })


@app.route("/api/fabric-global-backgrounds", methods=["GET", "POST"])
@admin_required
def fabric_global_backgrounds():
    prefs = load_admin_prefs()
    backgrounds = sanitize_global_backgrounds_list(prefs.get("fabric_global_backgrounds", []))

    if request.method == "GET":
        return jsonify({"backgrounds": backgrounds})

    payload = request.get_json(silent=True) or {}
    background_url = str(payload.get("url") or "").strip()
    if not background_url:
        return jsonify({"error": "Background URL is required"}), 400
    orientation = str(payload.get("orientation") or "landscape").strip().lower()
    if orientation not in ("landscape", "portrait"):
        orientation = "landscape"

    target_key = normalize_url_lookup_key(background_url)
    existing = next((item for item in backgrounds if normalize_url_lookup_key(item.get("url")) == target_key), None)
    if existing:
        return jsonify({"status": "exists", "background": existing, "backgrounds": backgrounds})

    now_text = current_timestamp_display()
    institute = canonicalize_institute_name(payload.get("institute"))
    entry = {
        "id": "global-bg-" + datetime.now().strftime("%Y%m%d%H%M%S") + "-" + "".join(
            random.choices(string.ascii_lowercase + string.digits, k=4)
        ),
        "name": str(payload.get("name") or f"Global Background {len(backgrounds) + 1}")[:80],
        "url": background_url,
        "orientation": orientation,
        "created_at": now_text,
        "updated_at": now_text,
        "institute_name": institute or "",
    }
    backgrounds.insert(0, entry)
    prefs["fabric_global_backgrounds"] = backgrounds[:2000]
    save_admin_prefs(prefs)
    return jsonify({"status": "saved", "background": entry, "backgrounds": prefs["fabric_global_backgrounds"]})


@app.route("/api/fabric-global-backgrounds/<background_id>", methods=["DELETE"])
@admin_required
def delete_fabric_global_background(background_id):
    target_id = str(background_id or "").strip()
    if not target_id:
        return jsonify({"error": "Background id is required"}), 400

    prefs = load_admin_prefs()
    backgrounds = sanitize_global_backgrounds_list(prefs.get("fabric_global_backgrounds", []))
    target = next((item for item in backgrounds if str(item.get("id") or "").strip() == target_id), None)
    if not target:
        return jsonify({"error": "Background not found"}), 404

    remaining = [item for item in backgrounds if str(item.get("id") or "").strip() != target_id]
    prefs["fabric_global_backgrounds"] = remaining
    save_admin_prefs(prefs)

    cleanup_result = remove_global_background_from_all_settings(target.get("url"))
    delete_generated_asset_from_url(target.get("url"))
    delete_supabase_storage_url(target.get("url"))

    return jsonify({
        "status": "deleted",
        "deleted_id": target_id,
        "deleted_url": str(target.get("url") or "").strip(),
        "cleanup": cleanup_result,
        "backgrounds": remaining,
    })


@app.route("/api/fabric-shared-backgrounds")
@admin_required
def fabric_shared_backgrounds():
    limit = request.args.get("limit", 120)
    try:
        limit_value = max(1, min(int(limit), 400))
    except Exception:
        limit_value = 120

    backgrounds = []
    seen_urls = set()

    def register_item(item_id, url, name="", institute_slug="", institute_name="", object_path="", source=""):
        value = str(url or "").strip()
        if not value:
            return False
        key = normalize_url_lookup_key(value)
        if key in seen_urls:
            return False
        seen_urls.add(key)
        backgrounds.append({
            "id": str(item_id or "").strip(),
            "url": value,
            "object_path": str(object_path or "").strip(),
            "label": str(name or "").strip(),
            "institute_slug": str(institute_slug or "").strip(),
            "institute_name": str(institute_name or "").strip(),
            "source": str(source or "").strip(),
        })
        return True

    prefs = load_admin_prefs()
    for item in sanitize_global_backgrounds_list(prefs.get("fabric_global_backgrounds", [])):
        register_item(
            item.get("id"),
            item.get("url"),
            name=item.get("name"),
            institute_slug=make_storage_slug(item.get("institute_name")),
            institute_name=item.get("institute_name"),
            source="global_pool",
        )
        if len(backgrounds) >= limit_value:
            return jsonify({"backgrounds": backgrounds})

    institutes = set()
    requested_institute = canonicalize_institute_name(request.args.get("institute"))
    if requested_institute:
        institutes.add(requested_institute)
    institutes.update(list_known_institutes_from_settings())
    try:
        for batch in list_all_supabase_batches():
            inst = canonicalize_institute_name(batch.get("institute_name"))
            if inst:
                institutes.add(inst)
    except Exception:
        pass

    for inst in sorted(institutes):
        settings = load_settings(inst)
        for item in sanitize_print_backgrounds_list(settings.get("print_backgrounds", [])):
            if register_item(
                item.get("id"),
                item.get("url"),
                name=item.get("name"),
                institute_slug=make_storage_slug(inst),
                institute_name=inst,
                source="institute_collection",
            ):
                if len(backgrounds) >= limit_value:
                    return jsonify({"backgrounds": backgrounds})

    return jsonify({"backgrounds": backgrounds})


@app.route("/api/admin/institutes")
@admin_required
def admin_institutes_api():
    institutes = set(list_known_institutes_from_settings())
    try:
        for batch in list_all_supabase_batches():
            institute = canonicalize_institute_name(batch.get("institute_name"))
            if institute:
                institutes.add(institute)
    except Exception:
        pass
    return jsonify({
        "institutes": sorted(institutes),
        "facility_institutes": list(FACILITY_LOCATION_INSTITUTES),
        "facility_structure": build_facility_structure_payload(),
    })


def normalize_background_orientation(value):
    orientation = str(value or "landscape").strip().lower()
    return orientation if orientation in ("landscape", "portrait") else "landscape"


def build_background_library_entry(entry_id, name, url, orientation, extra=None):
    payload = {
        "id": str(entry_id or "").strip(),
        "name": str(name or "Background").strip()[:80] or "Background",
        "url": str(url or "").strip(),
        "orientation": normalize_background_orientation(orientation),
        "created_at": current_timestamp_display(),
        "updated_at": current_timestamp_display(),
    }
    if isinstance(extra, dict):
        payload.update(extra)
    return payload


@app.route("/api/background-library", methods=["GET", "POST"])
@admin_required
def background_library_api():
    if request.method == "GET":
        scope = str(request.args.get("scope") or "common").strip().lower()
        institute = canonicalize_institute_name(request.args.get("institute"))
        block = str(request.args.get("block") or "").strip()
        facility_sub_location = str(request.args.get("facility_sub_location") or "").strip()

        if scope == "common":
            prefs = load_admin_prefs()
            items = sanitize_global_backgrounds_list(prefs.get("fabric_global_backgrounds", []))
        elif scope == "institute":
            if not institute:
                return jsonify({"items": [], "scope": scope, "institute": ""})
            settings = load_settings(institute)
            items = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))
        elif scope == "office":
            if not institute or not block or not facility_sub_location:
                return jsonify({"items": [], "scope": scope, "institute": institute or "", "block": block, "facility_sub_location": facility_sub_location})
            settings = load_settings(institute)
            items = [
                item for item in sanitize_office_backgrounds_list(settings.get("office_backgrounds", []))
                if str(item.get("block") or "").strip() == block and str(item.get("facility_sub_location") or "").strip() == facility_sub_location
            ]
        else:
            return jsonify({"error": "Invalid scope"}), 400

        return jsonify({
            "scope": scope,
            "institute": institute or "",
            "block": block,
            "facility_sub_location": facility_sub_location,
            "items": items,
        })

    scope = str(request.form.get("scope") or "common").strip().lower()
    institute = canonicalize_institute_name(request.form.get("institute"))
    block = str(request.form.get("block") or "").strip()
    facility_sub_location = str(request.form.get("facility_sub_location") or "").strip()
    orientation = normalize_background_orientation(request.form.get("orientation"))
    name = (request.form.get("name") or "").strip()

    if "background" not in request.files:
        return jsonify({"error": "No background file provided"}), 400
    file_storage = request.files["background"]
    if not secure_filename(file_storage.filename):
        return jsonify({"error": "Invalid filename"}), 400

    if scope == "common":
        storage_scope_name = "Common Background Library"
    elif scope == "institute":
        if not institute:
            return jsonify({"error": "Institute is required"}), 400
        storage_scope_name = institute
    elif scope == "office":
        if not institute:
            return jsonify({"error": "Institute is required"}), 400
        if not block:
            return jsonify({"error": "Block is required"}), 400
        if not facility_sub_location:
            return jsonify({"error": "Facility Sub Location is required"}), 400
        storage_scope_name = institute
    else:
        return jsonify({"error": "Invalid scope"}), 400

    entry_id = f"bg-{scope}-" + datetime.now().strftime("%Y%m%d%H%M%S") + "-" + "".join(random.choices(string.ascii_lowercase + string.digits, k=4))
    entry_name = name or os.path.splitext(secure_filename(file_storage.filename))[0] or "Background"
    try:
        background_url, _ = save_print_background_image(file_storage, storage_scope_name, entry_id)
    except Exception:
        return jsonify({"error": "Unable to process background image"}), 400

    if scope == "common":
        prefs = load_admin_prefs()
        items = sanitize_global_backgrounds_list(prefs.get("fabric_global_backgrounds", []))
        items.insert(0, build_background_library_entry(entry_id, entry_name, background_url, orientation, {"institute_name": ""}))
        prefs["fabric_global_backgrounds"] = items[:2000]
        save_admin_prefs(prefs)
        saved_items = prefs["fabric_global_backgrounds"]
    elif scope == "institute":
        settings = load_settings(institute)
        items = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))
        items.insert(0, build_background_library_entry(entry_id, entry_name, background_url, orientation))
        settings["print_backgrounds"] = items[:80]
        save_settings(settings, institute)
        saved_items = settings["print_backgrounds"]
    else:
        settings = load_settings(institute)
        items = sanitize_office_backgrounds_list(settings.get("office_backgrounds", []))
        items.insert(0, build_background_library_entry(entry_id, entry_name, background_url, orientation, {
            "block": block,
            "facility_sub_location": facility_sub_location,
        }))
        settings["office_backgrounds"] = items[:200]
        save_settings(settings, institute)
        saved_items = [
            item for item in settings["office_backgrounds"]
            if str(item.get("block") or "").strip() == block and str(item.get("facility_sub_location") or "").strip() == facility_sub_location
        ]

    return jsonify({
        "status": "saved",
        "scope": scope,
        "institute": institute or "",
        "block": block,
        "facility_sub_location": facility_sub_location,
        "items": saved_items,
    })


@app.route("/api/background-library/<background_id>", methods=["DELETE"])
@admin_required
def delete_background_library_item(background_id):
    item_id = str(background_id or "").strip()
    if not item_id:
        return jsonify({"error": "Background id is required"}), 400

    scope = str(request.args.get("scope") or "common").strip().lower()
    institute = canonicalize_institute_name(request.args.get("institute"))
    block = str(request.args.get("block") or "").strip()
    facility_sub_location = str(request.args.get("facility_sub_location") or "").strip()

    target = None
    remaining = []
    deleted_url = ""

    if scope == "common":
        prefs = load_admin_prefs()
        items = sanitize_global_backgrounds_list(prefs.get("fabric_global_backgrounds", []))
        target = next((item for item in items if str(item.get("id") or "").strip() == item_id), None)
        if not target:
            return jsonify({"error": "Background not found"}), 404
        remaining = [item for item in items if str(item.get("id") or "").strip() != item_id]
        prefs["fabric_global_backgrounds"] = remaining
        save_admin_prefs(prefs)
    elif scope == "institute":
        if not institute:
            return jsonify({"error": "Institute is required"}), 400
        settings = load_settings(institute)
        items = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))
        target = next((item for item in items if str(item.get("id") or "").strip() == item_id), None)
        if not target:
            return jsonify({"error": "Background not found"}), 404
        remaining = [item for item in items if str(item.get("id") or "").strip() != item_id]
        settings["print_backgrounds"] = remaining
        if str(settings.get("print_background_active_id") or "").strip() == item_id:
            settings["print_background_active_id"] = remaining[0].get("id", "") if remaining else ""
            settings["background_url"] = remaining[0].get("url", "") if remaining else ""
            settings["background_drive_id"] = ""
        save_settings(settings, institute)
    elif scope == "office":
        if not institute:
            return jsonify({"error": "Institute is required"}), 400
        settings = load_settings(institute)
        items = sanitize_office_backgrounds_list(settings.get("office_backgrounds", []))
        target = next((item for item in items if str(item.get("id") or "").strip() == item_id), None)
        if not target:
            return jsonify({"error": "Background not found"}), 404
        remaining_all = [item for item in items if str(item.get("id") or "").strip() != item_id]
        settings["office_backgrounds"] = remaining_all
        save_settings(settings, institute)
        remaining = [
            item for item in remaining_all
            if (not block or str(item.get("block") or "").strip() == block)
            and (not facility_sub_location or str(item.get("facility_sub_location") or "").strip() == facility_sub_location)
        ]
    else:
        return jsonify({"error": "Invalid scope"}), 400

    deleted_url = str((target or {}).get("url") or "").strip()
    if deleted_url:
        try:
            delete_generated_asset_from_url(deleted_url)
        except Exception:
            pass
        try:
            delete_supabase_storage_url(deleted_url)
        except Exception:
            pass

    return jsonify({
        "status": "deleted",
        "scope": scope,
        "institute": institute or "",
        "block": block,
        "facility_sub_location": facility_sub_location,
        "deleted_id": item_id,
        "deleted_url": deleted_url,
        "items": remaining,
    })


@app.route("/api/fabric-batch-background", methods=["GET", "POST"])
@admin_required
def fabric_batch_background():
    if request.method == "GET":
        institute = canonicalize_institute_name(request.args.get("institute"))
        batch_id = str(request.args.get("batch_id") or "").strip()
    else:
        payload = request.get_json(silent=True) or {}
        institute = canonicalize_institute_name(payload.get("institute"))
        batch_id = str(payload.get("batch_id") or "").strip()
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    if not batch_id:
        return jsonify({"error": "Batch is required"}), 400

    settings = load_settings(institute)
    bindings = sanitize_fabric_batch_background_bindings(settings.get("fabric_batch_background_bindings", {}))

    if request.method == "GET":
        return jsonify({
            "institute": institute,
            "batch_id": batch_id,
            "background_url": str(bindings.get(batch_id) or "").strip(),
            "bindings_count": len(bindings),
        })

    payload = request.get_json(silent=True) or {}
    background_url = str(payload.get("background_url") or "").strip()
    if background_url:
        bindings[batch_id] = background_url
    else:
        bindings.pop(batch_id, None)
    settings["fabric_batch_background_bindings"] = bindings
    save_settings(settings, institute)
    return jsonify({
        "status": "saved",
        "institute": institute,
        "batch_id": batch_id,
        "background_url": str(bindings.get(batch_id) or "").strip(),
        "bindings_count": len(bindings),
    })


@app.route("/admin/login", methods=["POST"])
def admin_login():
    password = request.form.get("password", "")
    if password == app.config["ADMIN_PANEL_PASSWORD"]:
        session["admin_authenticated"] = True
        log_audit_event("admin_login", success=True, ip=get_client_ip())
        return redirect(url_for("admin"))
    log_audit_event("admin_login", success=False, ip=get_client_ip())
    return render_template("admin_login.html", error="Incorrect admin password"), 401


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    log_audit_event("admin_logout", ip=get_client_ip())
    session.pop("admin_authenticated", None)
    return redirect(url_for("admin"))


@app.route("/api/admin-activity-log")
@admin_required
def admin_activity_log():
    limit = request.args.get("limit", 200)
    per_day = request.args.get("per_day", 20)
    since_days = request.args.get("since_days", 3)
    source = (request.args.get("source") or "supabase").strip().lower()
    try:
        limit_value = max(1, min(int(limit), 1000))
    except Exception:
        limit_value = 200
    try:
        per_day_value = max(1, min(int(per_day), 100))
    except Exception:
        per_day_value = 20
    try:
        since_days_value = max(1, min(int(since_days), 30))
    except Exception:
        since_days_value = 3

    if source in ("supabase", "readable", "remote") and is_supabase_enabled():
        try:
            remote_data = load_readable_logs_from_supabase(
                since_days=since_days_value,
                per_day=per_day_value,
                limit=limit_value,
            )
            return jsonify({
                "audit_log": remote_data.get("audit_log", []),
                "day_wise_audit": remote_data.get("day_wise_audit", []),
                "app_log": remote_data.get("app_log", []),
                "since_days": since_days_value,
                "source": remote_data.get("source", "supabase_readable"),
                "files": {
                    "audit": "supabase://system-logs/readable/<date>/audit.jsonl",
                    "app": "supabase://system-logs/readable/<date>/app.log",
                },
            })
        except Exception:
            pass

    audit_entries = filter_audit_entries_since_days(
        parse_audit_lines(read_recent_log_entries(AUDIT_LOG_FILE, limit_value)),
        since_days_value,
    )
    day_wise = build_day_wise_audit(audit_entries, per_day=per_day_value)
    app_entries = filter_app_entries_since_days(read_recent_log_entries(APP_LOG_FILE, limit_value), since_days_value)

    return jsonify({
        "audit_log": audit_entries,
        "day_wise_audit": day_wise,
        "app_log": app_entries,
        "since_days": since_days_value,
        "source": "local_files",
        "files": {
            "audit": AUDIT_LOG_FILE,
            "app": APP_LOG_FILE,
        },
    })


@app.route("/api/admin-activity-log/download")
@admin_required
def admin_activity_log_download():
    limit = request.args.get("limit", 1000)
    per_day = request.args.get("per_day", 25)
    since_days = request.args.get("since_days", 3)
    try:
        limit_value = max(1, min(int(limit), 2000))
    except Exception:
        limit_value = 1000
    try:
        per_day_value = max(1, min(int(per_day), 200))
    except Exception:
        per_day_value = 25
    try:
        since_days_value = max(1, min(int(since_days), 30))
    except Exception:
        since_days_value = 3

    archive_bytes, summary = build_activity_logs_archive(limit=limit_value, per_day=per_day_value, since_days=since_days_value)
    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    filename = f"activity_logs_last_{since_days_value}d_{timestamp}.zip"

    backup_url = ""
    backup_error = ""
    if is_supabase_enabled():
        object_path = f"system-logs/activity/{datetime.utcnow().strftime('%Y-%m-%d')}/{filename}"
        try:
            backup_url = upload_bytes_to_supabase_storage(archive_bytes, object_path, "application/zip")
        except Exception as exc:
            backup_error = str(exc) or "Unable to backup logs to Supabase"
            app.logger.warning("Log backup upload failed", exc_info=True)

    log_audit_event(
        "admin_activity_log_download",
        filename=filename,
        limit=limit_value,
        per_day=per_day_value,
        since_days=since_days_value,
        total_audit_events=summary.get("total_audit_events", 0),
        backup_url=backup_url,
        backup_error=backup_error,
    )

    response = send_file(
        io.BytesIO(archive_bytes),
        mimetype="application/zip",
        as_attachment=True,
        download_name=filename,
    )
    if backup_url:
        response.headers["X-Log-Backup-Url"] = backup_url
    if backup_error:
        response.headers["X-Log-Backup-Error"] = backup_error
    return response


@app.route("/api/cron/daily-log-backup")
def cron_daily_log_backup():
    if not verify_cron_request():
        return jsonify({"error": "Unauthorized"}), 401

    if not is_supabase_enabled():
        log_audit_event("cron_daily_log_backup", success=False, error="Supabase is not configured")
        return jsonify({"error": "Supabase is not configured"}), 500

    try:
        summary = sync_readable_logs_to_supabase(since_days=3, limit=5000, per_day=50)
    except Exception as exc:
        message = str(exc) or "Unable to sync daily readable logs"
        log_audit_event("cron_daily_log_backup", success=False, error=message)
        return jsonify({"error": message}), 500

    log_audit_event(
        "cron_daily_log_backup",
        success=True,
        mode="readable_jsonl",
        uploaded_days=summary.get("uploaded_days", []),
        total_audit_events=summary.get("total_audit_events", 0),
        total_app_lines=summary.get("total_app_lines", 0),
    )
    return jsonify({
        "status": "ok",
        "mode": "readable_jsonl",
        "uploaded_days": summary.get("uploaded_days", []),
        "total_audit_events": summary.get("total_audit_events", 0),
        "total_app_lines": summary.get("total_app_lines", 0),
    })


@app.route("/api/settings")
def get_settings():
    institute = canonicalize_institute_name(request.args.get("institute"))
    settings = load_settings(institute)
    return jsonify({
        "background_url": settings.get("background_url", ""),
        "signature_url": settings.get("signature_url", ""),
        "facility_custom_sub_locations": sanitize_facility_custom_sub_locations(settings.get("facility_custom_sub_locations", {})),
        "institute": institute
    })


@app.route("/api/certificate-settings")
def get_certificate_settings():
    institute = canonicalize_institute_name(request.args.get("institute"))
    settings = load_settings(institute)
    return jsonify({
        "background_url": settings.get("certificate_background_url", ""),
        "institute": institute
    })


@app.route("/api/facility-sub-locations", methods=["POST"])
def save_facility_sub_locations():
    payload = request.get_json(silent=True) or {}
    institute = canonicalize_institute_name(payload.get("institute"))
    block = str(payload.get("block") or "").strip()
    value = str(payload.get("value") or "").strip()
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    if not block:
        return jsonify({"error": "Block is required"}), 400
    if not value or value == "Add New":
        return jsonify({"error": "Village / Town is required"}), 400

    settings = load_settings(institute)
    custom_map = sanitize_facility_custom_sub_locations(settings.get("facility_custom_sub_locations", {}))
    current_items = list(custom_map.get(block, []))
    if value not in current_items:
        current_items.append(value)
    custom_map[block] = sorted({str(item).strip() for item in current_items if str(item).strip() and str(item).strip() != "Add New"})
    settings["facility_custom_sub_locations"] = custom_map
    save_settings(settings, institute)
    return jsonify({
        "status": "saved",
        "institute": institute,
        "block": block,
        "items": custom_map.get(block, []),
    })


@app.route("/api/print-templates", methods=["GET", "POST"])
@admin_required
def print_templates_api():
    institute = canonicalize_institute_name(request.args.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400

    settings = load_settings(institute)
    templates = sanitize_print_templates_list(settings.get("print_templates", []))
    if templates != settings.get("print_templates", []):
        settings["print_templates"] = templates
        save_settings(settings, institute)

    if request.method == "GET":
        return jsonify({"templates": templates, "institute": institute})

    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Template name is required"}), 400
    config = sanitize_print_template_config(data.get("config"))
    if not config:
        return jsonify({"error": "Template config is required"}), 400

    template_id = "tpl-" + datetime.now().strftime("%Y%m%d%H%M%S") + "-" + "".join(
        random.choices(string.ascii_lowercase + string.digits, k=4)
    )
    now_text = current_timestamp_display()
    template = {
        "id": template_id,
        "name": name[:80],
        "config": config,
        "created_at": now_text,
        "updated_at": now_text,
    }
    templates.insert(0, template)
    settings["print_templates"] = templates[:60]
    save_settings(settings, institute)
    return jsonify({"status": "saved", "template": template, "templates": settings["print_templates"], "institute": institute})


@app.route("/api/print-templates/<template_id>", methods=["DELETE"])
@admin_required
def delete_print_template(template_id):
    institute = canonicalize_institute_name(request.args.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    settings = load_settings(institute)
    templates = sanitize_print_templates_list(settings.get("print_templates", []))
    filtered = [template for template in templates if template.get("id") != template_id]
    if len(filtered) == len(templates):
        return jsonify({"error": "Template not found"}), 404
    settings["print_templates"] = filtered
    save_settings(settings, institute)
    return jsonify({"status": "deleted", "template_id": template_id, "templates": filtered, "institute": institute})


@app.route("/api/drive-storage-status")
@admin_required
def drive_storage_status():
    institute = canonicalize_institute_name(request.args.get("institute"))
    return jsonify(build_drive_storage_status(institute))

@app.route("/uploads/<filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)

@app.route("/generated-assets/<filename>")
def generated_asset(filename):
    return send_from_directory(ASSET_DIR, filename)

@app.route("/api/upload-photo", methods=["POST"])
def upload_photo():
    if "photo" not in request.files:
        return jsonify({"error": "No file"}), 400
    try:
        institute = canonicalize_institute_name(request.form.get("institute"))
        serial = (request.form.get("serial_no") or "").strip() or gen_serial(institute)
        photo_url, photo_drive_id = attach_photo_to_serial(
            request.files["photo"],
            serial,
            institute=institute,
        )
        return jsonify({"serial_no": serial, "photo_url": photo_url, "photo_drive_id": photo_drive_id})
    except ValueError as exc:
        return jsonify({"error": str(exc) or "Unable to process photo"}), 400
    except RuntimeError as exc:
        app.logger.warning("Photo upload failed", exc_info=True)
        return jsonify({"error": str(exc) or "Photo upload failed"}), 500
    except Exception:
        app.logger.exception("Unexpected photo upload failure")
        return jsonify({"error": "Unexpected photo upload failure"}), 500


def attach_photo_to_serial(file_storage, serial, institute=None):
    if not serial:
        raise ValueError("Serial number is required")
    filename = f"{serial}.jpg"
    raw_path = os.path.join(UPLOAD_DIR, f"raw_{filename}")
    final_path = os.path.join(UPLOAD_DIR, filename)
    file_storage.save(raw_path)
    try:
        autocrop_passport(raw_path, final_path)
        os.remove(raw_path)
    except Exception:
        os.rename(raw_path, final_path)
    photo_url = f"/uploads/{filename}"
    photo_drive_id = None
    if is_supabase_enabled():
        institute = canonicalize_institute_name(institute or request.form.get("institute"))
        storage_path = build_photo_storage_path(institute, serial, file_storage.filename)
        try:
            with open(final_path, "rb") as photo_file:
                photo_url = upload_bytes_to_supabase_storage(photo_file.read(), storage_path, "image/jpeg")
        except Exception as exc:
            raise RuntimeError(f"Cloud photo upload failed for {serial}: {exc}") from exc
    elif is_drive_enabled():
        with open(final_path, "rb") as photo_file:
            photo_drive_id, drive_url = upload_bytes_to_drive(photo_file.read(), filename, "image/jpeg", "photos")
            if drive_url:
                photo_url = drive_url
    return photo_url, photo_drive_id


@app.route("/api/print-backgrounds", methods=["GET", "POST"])
@admin_required
def print_backgrounds_api():
    institute = canonicalize_institute_name(request.args.get("institute") or request.form.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400

    settings = load_settings(institute)
    ensure_print_background_state(settings)

    if request.method == "GET":
        return jsonify({
            "institute": institute,
            "backgrounds": settings.get("print_backgrounds", []),
            "active_background_id": settings.get("print_background_active_id", ""),
            "background_url": settings.get("background_url", ""),
        })

    if "background" not in request.files:
        return jsonify({"error": "No file"}), 400
    file_storage = request.files["background"]
    if not secure_filename(file_storage.filename):
        return jsonify({"error": "Invalid filename"}), 400

    background_name = (request.form.get("name") or "").strip() or os.path.splitext(secure_filename(file_storage.filename))[0] or "Background"
    orientation = str(request.form.get("orientation") or "landscape").strip().lower()
    if orientation not in ("landscape", "portrait"):
        orientation = "landscape"
    background_id = "bg-" + datetime.now().strftime("%Y%m%d%H%M%S") + "-" + "".join(
        random.choices(string.ascii_lowercase + string.digits, k=4)
    )

    try:
        background_url, _ = save_print_background_image(file_storage, institute, background_id)
    except Exception:
        return jsonify({"error": "Unable to process background image"}), 400

    now_text = current_timestamp_display()
    entry = {
        "id": background_id,
        "name": background_name[:80],
        "url": background_url,
        "orientation": orientation,
        "created_at": now_text,
        "updated_at": now_text,
    }
    backgrounds = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))
    backgrounds.insert(0, entry)
    settings["print_backgrounds"] = backgrounds[:80]
    settings["print_background_active_id"] = background_id
    settings["background_url"] = background_url
    settings["background_drive_id"] = ""
    save_settings(settings, institute)
    return jsonify({
        "status": "saved",
        "institute": institute,
        "background": entry,
        "backgrounds": settings.get("print_backgrounds", []),
        "active_background_id": settings.get("print_background_active_id", ""),
        "background_url": settings.get("background_url", ""),
    })


@app.route("/api/print-backgrounds/<background_id>/activate", methods=["POST"])
@admin_required
def activate_print_background(background_id):
    institute = canonicalize_institute_name(request.args.get("institute") or (request.json or {}).get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    settings = load_settings(institute)
    backgrounds = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))
    target = next((item for item in backgrounds if item.get("id") == background_id), None)
    if not target:
        return jsonify({"error": "Background not found"}), 404
    settings["print_backgrounds"] = backgrounds
    settings["print_background_active_id"] = target.get("id", "")
    settings["background_url"] = target.get("url", "")
    settings["background_drive_id"] = ""
    save_settings(settings, institute)
    return jsonify({
        "status": "activated",
        "institute": institute,
        "active_background_id": settings.get("print_background_active_id", ""),
        "background_url": settings.get("background_url", ""),
        "backgrounds": settings.get("print_backgrounds", []),
    })


@app.route("/api/print-backgrounds/link", methods=["POST"])
@admin_required
def link_print_background():
    payload = request.get_json(silent=True) or {}
    institute = canonicalize_institute_name(payload.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    background_url = str(payload.get("url") or "").strip()
    if not background_url:
        return jsonify({"error": "Background URL is required"}), 400

    settings = load_settings(institute)
    ensure_print_background_state(settings)
    backgrounds = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))

    def normalize_bg_key(value):
        return str(value or "").strip().split("?", 1)[0].lower()

    target_key = normalize_bg_key(background_url)
    existing = next((item for item in backgrounds if normalize_bg_key(item.get("url")) == target_key), None)
    now_text = current_timestamp_display()
    orientation = str(payload.get("orientation") or "landscape").strip().lower()
    if orientation not in ("landscape", "portrait"):
        orientation = "landscape"
    if existing:
        settings["print_background_active_id"] = existing.get("id", "")
        settings["background_url"] = existing.get("url", "")
        save_settings(settings, institute)
        return jsonify({
            "status": "linked",
            "institute": institute,
            "background": existing,
            "backgrounds": settings.get("print_backgrounds", []),
            "active_background_id": settings.get("print_background_active_id", ""),
            "background_url": settings.get("background_url", ""),
        })

    name = str(payload.get("name") or "").strip() or f"Background {len(backgrounds) + 1}"
    background_id = "bg-link-" + datetime.now().strftime("%Y%m%d%H%M%S") + "-" + "".join(
        random.choices(string.ascii_lowercase + string.digits, k=4)
    )
    entry = {
        "id": background_id,
        "name": name[:80],
        "url": background_url,
        "orientation": orientation,
        "created_at": now_text,
        "updated_at": now_text,
    }
    backgrounds.insert(0, entry)
    settings["print_backgrounds"] = backgrounds[:80]
    settings["print_background_active_id"] = background_id
    settings["background_url"] = background_url
    save_settings(settings, institute)
    return jsonify({
        "status": "linked",
        "institute": institute,
        "background": entry,
        "backgrounds": settings.get("print_backgrounds", []),
        "active_background_id": settings.get("print_background_active_id", ""),
        "background_url": settings.get("background_url", ""),
    })


@app.route("/api/print-backgrounds/<background_id>", methods=["DELETE"])
@admin_required
def delete_print_background(background_id):
    institute = canonicalize_institute_name(request.args.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    settings = load_settings(institute)
    backgrounds = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))
    target = next((item for item in backgrounds if item.get("id") == background_id), None)
    if not target:
        return jsonify({"error": "Background not found"}), 404
    remaining = [item for item in backgrounds if item.get("id") != background_id]
    settings["print_backgrounds"] = remaining
    if settings.get("print_background_active_id") == background_id:
        if remaining:
            settings["print_background_active_id"] = remaining[0].get("id", "")
            settings["background_url"] = remaining[0].get("url", "")
        else:
            settings["print_background_active_id"] = ""
            settings["background_url"] = ""
            settings["background_drive_id"] = ""
    save_settings(settings, institute)
    delete_generated_asset_from_url(target.get("url"))
    return jsonify({
        "status": "deleted",
        "institute": institute,
        "deleted_background_id": background_id,
        "active_background_id": settings.get("print_background_active_id", ""),
        "background_url": settings.get("background_url", ""),
        "backgrounds": settings.get("print_backgrounds", []),
    })


@app.route("/api/upload-background", methods=["POST"])
@admin_required
def upload_background():
    if "background" not in request.files:
        return jsonify({"error": "No file"}), 400
    institute = canonicalize_institute_name(request.form.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    file_storage = request.files["background"]
    if not secure_filename(file_storage.filename):
        return jsonify({"error": "Invalid filename"}), 400
    try:
        background_url, background_drive_id = save_background_image(file_storage, institute)
    except Exception:
        return jsonify({"error": "Unable to process background image"}), 400
    settings = load_settings(institute)
    old_drive_id = settings.get("background_drive_id")
    if old_drive_id and old_drive_id != background_drive_id:
        delete_drive_file(old_drive_id)
    settings["background_url"] = background_url
    settings["background_drive_id"] = background_drive_id or ""
    backgrounds = sanitize_print_backgrounds_list(settings.get("print_backgrounds", []))
    primary_entry = next((item for item in backgrounds if item.get("id") == "primary"), None)
    now_text = current_timestamp_display()
    if primary_entry:
        primary_entry["url"] = background_url
        primary_entry["updated_at"] = now_text
    else:
        backgrounds.insert(0, {
            "id": "primary",
            "name": "Primary Background",
            "url": background_url,
            "created_at": now_text,
            "updated_at": now_text,
        })
    settings["print_backgrounds"] = backgrounds[:80]
    settings["print_background_active_id"] = "primary"
    save_settings(settings, institute)
    return jsonify({"status": "saved", "background_url": background_url, "institute": institute})


@app.route("/api/upload-certificate-background", methods=["POST"])
@admin_required
def upload_certificate_background():
    if "background" not in request.files:
        return jsonify({"error": "No file"}), 400
    institute = canonicalize_institute_name(request.form.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    file_storage = request.files["background"]
    if not secure_filename(file_storage.filename):
        return jsonify({"error": "Invalid filename"}), 400
    try:
        background_url, background_drive_id = save_certificate_background_image(file_storage, institute)
    except Exception:
        return jsonify({"error": "Unable to process certificate background image"}), 400

    settings = load_settings(institute)
    old_drive_id = settings.get("certificate_background_drive_id")
    if old_drive_id and old_drive_id != background_drive_id:
        delete_drive_file(old_drive_id)
    settings["certificate_background_url"] = background_url
    settings["certificate_background_drive_id"] = background_drive_id or ""
    save_settings(settings, institute)
    return jsonify({"status": "saved", "background_url": background_url, "institute": institute})


@app.route("/api/upload-signature", methods=["POST"])
def upload_signature():
    institute = canonicalize_institute_name(request.form.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    if request.form.get("password", "") != app.config["SIGNATURE_UPLOAD_PASSWORD"]:
        return jsonify({"error": "Incorrect password"}), 403
    if "signature" not in request.files:
        return jsonify({"error": "No file"}), 400
    file_storage = request.files["signature"]
    if not secure_filename(file_storage.filename):
        return jsonify({"error": "Invalid filename"}), 400
    try:
        signature_url, signature_drive_id = save_signature_image(file_storage, institute)
    except Exception:
        return jsonify({"error": "Unable to process signature image"}), 400
    settings = load_settings(institute)
    old_signature_drive_id = settings.get("signature_drive_id")
    if old_signature_drive_id and old_signature_drive_id != signature_drive_id:
        delete_drive_file(old_signature_drive_id)
    settings["signature_url"] = signature_url
    settings["signature_drive_id"] = signature_drive_id or ""
    save_settings(settings, institute)
    return jsonify({"status": "saved", "signature_url": signature_url, "institute": institute})


@app.route("/api/delete-background", methods=["DELETE"])
@admin_required
def delete_background():
    institute = canonicalize_institute_name(request.args.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400

    settings = load_settings(institute)
    old_drive_id = settings.get("background_drive_id")
    old_url = settings.get("background_url", "")
    if old_drive_id:
        delete_drive_file(old_drive_id)
    settings["background_url"] = ""
    settings["background_drive_id"] = ""
    settings["print_background_active_id"] = ""
    settings["print_backgrounds"] = []
    delete_local_file_if_exists(os.path.join(ASSET_DIR, f"card_background_{make_asset_slug(institute)}.jpg"))
    save_settings(settings, institute)
    return jsonify({"status": "deleted", "background_url": old_url, "institute": institute})


@app.route("/api/delete-certificate-background", methods=["DELETE"])
@admin_required
def delete_certificate_background():
    institute = canonicalize_institute_name(request.args.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400

    settings = load_settings(institute)
    old_drive_id = settings.get("certificate_background_drive_id")
    old_url = settings.get("certificate_background_url", "")
    if old_drive_id:
        delete_drive_file(old_drive_id)
    settings["certificate_background_url"] = ""
    settings["certificate_background_drive_id"] = ""
    delete_local_file_if_exists(os.path.join(ASSET_DIR, f"certificate_background_{make_asset_slug(institute)}.jpg"))
    save_settings(settings, institute)
    return jsonify({"status": "deleted", "background_url": old_url, "institute": institute})


@app.route("/api/delete-signature", methods=["DELETE"])
@admin_required
def delete_signature():
    institute = canonicalize_institute_name(request.args.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    settings = load_settings(institute)
    old_drive_id = settings.get("signature_drive_id")
    if old_drive_id:
        delete_drive_file(old_drive_id)
    settings["signature_url"] = ""
    settings["signature_drive_id"] = ""
    delete_local_file_if_exists(os.path.join(ASSET_DIR, f"hod_signature_{make_asset_slug(institute)}.png"))
    save_settings(settings, institute)
    return jsonify({"status": "deleted", "institute": institute})

@app.route("/api/submit", methods=["POST"])
def submit():
    data = request.json
    data["institute_name"] = canonicalize_institute_name(data.get("institute_name"))
    institute = data.get("institute_name")
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    normalize_record_dates(data)
    data["saved_at"] = current_timestamp_display()
    records = load_records(institute)
    # Check for duplicate serial
    for r in records:
        if r.get("serial_no") == data.get("serial_no"):
            submitted_at = r.get("submitted_at")
            batch_total_cards = r.get("batch_total_cards")
            r.update(data)
            if submitted_at:
                r["submitted_at"] = submitted_at
            if batch_total_cards:
                r["batch_total_cards"] = batch_total_cards
            save_records(records, institute)
            return jsonify({"status": "updated"})
    data["submitted_at"] = None
    records.append(data)
    save_records(records, institute)
    return jsonify({"status": "saved"})


@app.route("/api/certificates", methods=["GET", "POST"])
def certificates_api():
    if request.method == "POST":
        data = request.json or {}
        if not (data.get("recipient_name") or "").strip():
            return jsonify({"error": "Recipient name is required"}), 400
        institute_name = (data.get("institute_name") or "").strip()
        if not institute_name:
            return jsonify({"error": "Institute is required"}), 400
        data["institute_name"] = canonicalize_institute_name(institute_name)

        certificates = load_certificates(data["institute_name"])
        certificate_no = (data.get("certificate_no") or "").strip() or gen_certificate_no()
        data["certificate_no"] = certificate_no
        normalize_record_dates(data)
        data["saved_at"] = current_timestamp_display()

        for certificate in certificates:
            if certificate.get("certificate_no") == certificate_no:
                certificate.update(data)
                save_certificates(certificates, data["institute_name"])
                return jsonify({"status": "updated", "certificate_no": certificate_no})

        certificates.append(data)
        save_certificates(certificates, data["institute_name"])
        return jsonify({"status": "saved", "certificate_no": certificate_no})

    institute = canonicalize_institute_name(request.args.get("institute"))
    certificates = load_certificates(institute)
    return jsonify({"certificates": certificates, "institute": institute})


@app.route("/api/certificates/<certificate_no>", methods=["DELETE"])
@admin_required
def delete_certificate(certificate_no):
    certificate_to_delete = next((c for c in load_certificates() if c.get("certificate_no") == certificate_no), None)
    if not certificate_to_delete:
        return jsonify({"error": "Certificate not found"}), 404
    institute = canonicalize_institute_name(certificate_to_delete.get("institute_name"))
    certificates = load_certificates(institute)
    certificates = [c for c in certificates if c.get("certificate_no") != certificate_no]
    save_certificates(certificates, institute)

    if certificate_to_delete.get("photo_drive_id"):
        delete_drive_file(certificate_to_delete.get("photo_drive_id"))
    delete_uploaded_file_from_url(certificate_to_delete.get("photo_url"))

    return jsonify({"status": "deleted", "certificate_no": certificate_no})


@app.route("/api/batch-summary")
def batch_summary():
    records, institute = get_filtered_records()
    summary = summarize_batch(records)
    summary["institute"] = institute
    return jsonify(summary)


@app.route("/api/batch-records")
def batch_records():
    records, institute = get_filtered_records()
    return jsonify({"records": records, "institute": institute})


@app.route("/api/batches")
@admin_required
def get_batches():
    institute = canonicalize_institute_name(request.args.get("institute"))
    if not institute:
        return jsonify({"batches": [], "institute": ""})
    if not is_supabase_enabled():
        return jsonify({"batches": [], "institute": institute, "warning": "Supabase is not configured"})
    try:
        batches = list_supabase_batches(institute)
    except Exception as exc:
        return jsonify({"error": str(exc) or "Unable to load batches"}), 500
    return jsonify({"batches": batches, "institute": institute})


@app.route("/api/batch-overview")
@admin_required
def batch_overview():
    institute = canonicalize_institute_name(request.args.get("institute"))
    if not is_supabase_enabled():
        return jsonify({"batches": [], "warning": "Supabase is not configured"})
    try:
        batches = list_supabase_batches(institute) if institute else list_all_supabase_batches()
        batches = enrich_batches_for_overview(batches)
    except Exception as exc:
        return jsonify({"error": str(exc) or "Unable to load batch overview"}), 500
    return jsonify({"batches": batches, "institute": institute or ""})


@app.route("/api/supabase-storage-status")
@admin_required
def supabase_storage_status():
    institute = canonicalize_institute_name(request.args.get("institute"))
    try:
        run_supabase_auto_cleanup()
        return jsonify(build_supabase_storage_status(institute or None))
    except Exception as exc:
        return jsonify({"error": str(exc) or "Unable to load Supabase status"}), 500


@app.route("/api/admin-storage-policy")
@admin_required
def get_admin_storage_policy():
    return jsonify(load_admin_prefs())


@app.route("/api/admin-storage-policy", methods=["POST"])
@admin_required
def save_admin_storage_policy():
    payload = request.json or {}
    prefs = load_admin_prefs()
    if "storage_quota_mb" in payload:
        try:
            prefs["storage_quota_mb"] = max(1, int(payload.get("storage_quota_mb") or 1024))
        except Exception:
            return jsonify({"error": "Storage quota must be a number"}), 400
    if "auto_delete_days" in payload:
        try:
            prefs["auto_delete_days"] = max(1, int(payload.get("auto_delete_days") or 15))
        except Exception:
            return jsonify({"error": "Auto delete days must be a number"}), 400
    if "auto_delete_enabled" in payload:
        prefs["auto_delete_enabled"] = bool(payload.get("auto_delete_enabled"))
    save_admin_prefs(prefs)
    cleanup_result = run_supabase_auto_cleanup(force=True)
    return jsonify({
        "status": "saved",
        "policy": load_admin_prefs(),
        "cleanup": cleanup_result,
    })


@app.route("/api/supabase-storage-test", methods=["POST"])
@admin_required
def supabase_storage_test():
    institute = canonicalize_institute_name((request.json or {}).get("institute"))
    if not is_supabase_enabled():
        return jsonify({"error": "Supabase is not configured"}), 500
    try:
        ensure_supabase_bucket(app.config["SUPABASE_PHOTOS_BUCKET"])
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
        probe_path = f"system-tests/{make_storage_slug(institute or 'default')}/probe-{timestamp}.txt"
        probe_text = f"Supabase storage test {timestamp}".encode("utf-8")
        probe_url = upload_bytes_to_supabase_storage(probe_text, probe_path, "text/plain")
        delete_supabase_storage_url(probe_url)
        return jsonify({
            "status": "ok",
            "project_url": app.config["SUPABASE_URL"],
            "bucket_name": app.config["SUPABASE_PHOTOS_BUCKET"],
            "probe_path": probe_path,
            "probe_url": probe_url,
            "message": "Storage test completed successfully",
        })
    except Exception as exc:
        return jsonify({"error": str(exc) or "Storage test failed"}), 500


@app.route("/api/batches/<batch_id>", methods=["DELETE"])
@admin_required
def delete_batch(batch_id):
    institute = canonicalize_institute_name(request.args.get("institute"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    if not batch_id:
        return jsonify({"error": "Batch is required"}), 400
    if not is_supabase_enabled():
        return jsonify({"error": "Supabase is not configured"}), 500

    try:
        result = delete_supabase_batch_data(institute, batch_id)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 404
    except Exception as exc:
        return jsonify({"error": str(exc) or "Unable to delete batch"}), 500

    return jsonify(result)


@app.route("/api/batches/merge", methods=["POST"])
@admin_required
def merge_batches():
    payload = request.json or {}
    institute = canonicalize_institute_name(payload.get("institute"))
    batch_ids = payload.get("batch_ids") or []
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    if not isinstance(batch_ids, list):
        return jsonify({"error": "Batch IDs must be a list"}), 400
    try:
        result = merge_supabase_batches(institute, batch_ids)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": str(exc) or "Unable to merge batches"}), 500
    return jsonify(result)


@app.route("/api/submit-batch", methods=["POST"])
def submit_batch():
    payload = request.json or {}
    institute = canonicalize_institute_name(payload.get("institute_name"))
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    records = payload.get("records") or []
    if not isinstance(records, list) or not records:
        return jsonify({"error": "Batch records are required"}), 400
    if not is_supabase_enabled():
        return jsonify({"error": "Supabase is not configured"}), 500
    try:
        result = create_supabase_batch(institute, records, payload.get("batch_name"))
    except Exception as exc:
        return jsonify({"error": str(exc) or "Batch submit failed"}), 500
    return jsonify({"status": "submitted", **result})


@app.route("/api/retrieve-card")
def retrieve_card():
    institute = canonicalize_institute_name(request.args.get("institute"))
    serial_lookup = (request.args.get("serial_no") or "").strip()
    query_lookup = (request.args.get("query") or "").strip()
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    if not serial_lookup and not query_lookup:
        return jsonify({"error": "ID card number or name is required"}), 400

    lookup_value = serial_lookup or query_lookup
    if is_supabase_enabled():
        records = list_supabase_records(institute)
    else:
        records = load_records(institute)

    matches = find_records_by_serial_lookup(records, lookup_value)
    if not matches and query_lookup:
        matches = find_records_by_name_lookup(records, query_lookup)
    if len(matches) > 1:
        if query_lookup:
            return jsonify({"error": "More than one card matches this name. Please enter the full ID card number or a more specific name."}), 409
        return jsonify({"error": "More than one card matches this short ID. Please enter the full ID card number."}), 409
    record = matches[0] if matches else None
    if not record:
        return jsonify({"error": "Card not found"}), 404
    return jsonify({"record": record, "institute": institute, "serial_no": record.get("serial_no", lookup_value)})


@app.route("/api/update-card", methods=["POST"])
def update_card():
    payload = request.json or {}
    institute = canonicalize_institute_name(payload.get("institute_name"))
    serial = (payload.get("serial_no") or "").strip()
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    if not serial:
        return jsonify({"error": "ID card number is required"}), 400

    normalize_record_dates(payload)
    payload["institute_name"] = institute
    payload["serial_no"] = serial
    payload["saved_at"] = current_timestamp_display()

    if is_supabase_enabled():
        existing = get_supabase_record_by_serial(institute, serial)
        if not existing:
            return jsonify({"error": "Card not found"}), 404
        submitted_at = existing.get("submitted_at")
        batch_id = existing.get("batch_id")
        batch_name = existing.get("batch_name")
        existing.update(payload)
        existing["submitted_at"] = submitted_at
        if batch_id:
            existing["batch_id"] = batch_id
        if batch_name:
            existing["batch_name"] = batch_name
        supabase_request(
            "PATCH",
            "records",
            payload={
                "payload": existing,
                "saved_at": existing.get("saved_at", ""),
                "name": existing.get("name", ""),
                "profile_type": existing.get("profile_type", ""),
                "submitted_at": submitted_at or "",
            },
            query={"serial_no": f"eq.{serial}", "institute_name": f"eq.{institute}"},
        )
        return jsonify({"status": "updated", "record": existing, "serial_no": serial})

    records = load_records(institute)
    existing = next((rec for rec in records if rec.get("serial_no") == serial), None)
    if not existing:
        return jsonify({"error": "Card not found"}), 404
    submitted_at = existing.get("submitted_at")
    batch_total_cards = existing.get("batch_total_cards")
    existing.update(payload)
    existing["submitted_at"] = submitted_at
    if batch_total_cards:
        existing["batch_total_cards"] = batch_total_cards
    save_records(records, institute)
    return jsonify({"status": "updated", "record": existing, "serial_no": serial})

@app.route("/api/records")
@admin_required
def get_records():
    records, institute = get_filtered_records()
    return jsonify({"records": records, "institute": institute})


@app.route("/api/admin-attach-photo", methods=["POST"])
@admin_required
def admin_attach_photo():
    serial = (request.form.get("serial_no") or "").strip()
    institute = canonicalize_institute_name(request.form.get("institute_name"))
    if not serial:
        return jsonify({"error": "Serial number is required"}), 400
    if "photo" not in request.files:
        return jsonify({"error": "No photo file"}), 400

    if is_supabase_enabled():
        records = list_supabase_records(institute or None)
    else:
        records = load_records(institute) if institute else load_records()
    target_record = next((rec for rec in records if rec.get("serial_no") == serial), None)
    if not target_record and institute and is_supabase_enabled():
        target_record = next(
            (rec for rec in list_supabase_records() if rec.get("serial_no") == serial and canonicalize_institute_name(rec.get("institute_name")) == institute),
            None,
        )
    if not target_record:
        return jsonify({"error": "Record not found"}), 404
    institute = canonicalize_institute_name(target_record.get("institute_name"))

    try:
        result = save_photo_on_record(target_record, request.files["photo"], institute)
    except Exception:
        return jsonify({"error": "Unable to process photo"}), 400
    return jsonify({"status": "saved", **result})


@app.route("/api/admin-bulk-attach-photos", methods=["POST"])
@admin_required
def admin_bulk_attach_photos():
    institute = canonicalize_institute_name(request.form.get("institute_name"))
    batch_id = (request.form.get("batch_id") or "").strip()
    photo_files = [photo for photo in request.files.getlist("photos") if secure_filename(photo.filename)]
    if not institute:
        return jsonify({"error": "Institute is required"}), 400
    if not batch_id:
        return jsonify({"error": "Batch is required"}), 400
    if not photo_files:
        return jsonify({"error": "No photo files"}), 400

    records = list_supabase_records(institute, batch_id) if is_supabase_enabled() else [
        rec for rec in load_records(institute) if str(rec.get("batch_id", "")).strip() == batch_id
    ]
    if not records:
        return jsonify({"error": "No records found for this batch"}), 404

    sorted_records = sort_records_for_bulk_match(records)
    grouped_records = {}
    record_by_number = {}
    for record in sorted_records:
        name_key = normalize_match_text(record.get("name", ""))
        if name_key:
            grouped_records.setdefault(name_key, []).append(record)
        serial_info = extract_trailing_number_info(record.get("serial_no", ""))
        if serial_info["number"] is not None and serial_info["number"] not in record_by_number:
            record_by_number[serial_info["number"]] = record

    matched_serials = set()
    matched = 0
    unmatched_files = []

    for photo in photo_files:
        info = extract_trailing_number_info(photo.filename or "")
        target_record = None
        is_serial_number_file = not info["base"] and info["number"] is not None

        if not is_serial_number_file:
            candidates = grouped_records.get(info["base"], [])
            if candidates:
                if info["number"] is not None:
                    candidate_index = max(1, info["number"]) - 1
                    if candidate_index < len(candidates):
                        candidate = candidates[candidate_index]
                        if candidate.get("serial_no") not in matched_serials:
                            target_record = candidate
                if not target_record:
                    target_record = next((candidate for candidate in candidates if candidate.get("serial_no") not in matched_serials), None)

        if not target_record and is_serial_number_file and info["number"] is not None:
            candidate = record_by_number.get(info["number"])
            if candidate and candidate.get("serial_no") not in matched_serials:
                target_record = candidate

        if not target_record:
            unmatched_files.append(photo.filename)
            continue

        try:
            save_photo_on_record(target_record, photo, institute)
        except Exception:
            unmatched_files.append(photo.filename)
            continue
        matched += 1
        matched_serials.add(target_record.get("serial_no"))

    missing_records = [
        record.get("serial_no") or record.get("name") or "Unnamed"
        for record in records
        if not record.get("photo_url") and record.get("serial_no") not in matched_serials
    ]
    return jsonify({
        "status": "saved",
        "matched": matched,
        "unmatched_files": unmatched_files,
        "missing_records": missing_records,
        "batch_id": batch_id,
        "institute": institute,
    })


@app.route("/api/import-csv", methods=["POST"])
@admin_required
def import_csv():
    if "csv_file" not in request.files:
        return jsonify({"error": "No CSV file"}), 400

    uploaded = request.files["csv_file"]
    if not secure_filename(uploaded.filename):
        return jsonify({"error": "Invalid CSV filename"}), 400

    try:
        text = uploaded.read().decode("utf-8-sig")
    except UnicodeDecodeError:
        return jsonify({"error": "CSV must be UTF-8 encoded"}), 400

    rows = list(csv.DictReader(io.StringIO(text)))
    if not rows:
        return jsonify({"error": "CSV has no data rows"}), 400

    records_by_institute = {}
    imported = 0
    updated = 0

    field_map = {
        "serial_no": ["serial_no", "serial", "id"],
        "institute_name": ["institute_name", "institute"],
        "profile_type": ["profile_type", "type"],
        "name": ["name", "full_name"],
        "course": ["course"],
        "training_year": ["training_year", "year"],
        "batch_session": ["batch_session", "batch"],
        "father_name": ["father_name", "fathers_name", "father"],
        "aadhaar_no": ["aadhaar_no", "aadhaar", "aadhar_no", "aadhar"],
        "employee_id": ["employee_id", "emp_id"],
        "designation": ["designation"],
        "department": ["department", "dept", "section"],
        "dob": ["dob", "date_of_birth"],
        "contact": ["contact", "contact_no", "mobile"],
        "blood_group": ["blood_group", "blood"],
        "address": ["address"],
        "valid_upto": ["valid_upto", "validity"],
        "inserted_date": ["inserted_date", "issue_date", "issued_date"],
    }

    for row in rows:
        normalized = {}
        for target, aliases in field_map.items():
            value = ""
            for alias in aliases:
                if alias in row and str(row.get(alias) or "").strip():
                    value = str(row.get(alias) or "").strip()
                    break
            if value:
                normalized[target] = value

        normalized["institute_name"] = canonicalize_institute_name(normalized.get("institute_name"))
        normalize_record_dates(normalized)
        serial_no = (normalized.get("serial_no") or "").strip() or gen_serial(normalized.get("institute_name"))
        normalized["serial_no"] = serial_no
        if not normalized.get("name"):
            continue

        profile_type = (normalized.get("profile_type") or "").strip().lower()
        if profile_type not in {"student", "lecturer", "employee"}:
            profile_type = "student" if normalized.get("course") or normalized.get("training_year") or normalized.get("batch_session") else "employee"
        normalized["profile_type"] = profile_type

        institute = normalized.get("institute_name")
        if not institute:
            continue
        if institute not in records_by_institute:
            institute_records = load_records(institute)
            records_by_institute[institute] = {
                "records": institute_records,
                "record_map": {rec.get("serial_no"): rec for rec in institute_records if rec.get("serial_no")},
            }
        bucket = records_by_institute[institute]
        record_map = bucket["record_map"]

        existing = record_map.get(serial_no)
        if existing:
            preserved = {
                "submitted_at": existing.get("submitted_at"),
                "batch_total_cards": existing.get("batch_total_cards"),
                "photo_url": existing.get("photo_url", ""),
                "photo_drive_id": existing.get("photo_drive_id", ""),
            }
            existing.update(normalized)
            existing.update({k: v for k, v in preserved.items() if v is not None})
            existing["saved_at"] = current_timestamp_display()
            updated += 1
        else:
            normalized.setdefault("photo_url", "")
            normalized.setdefault("photo_drive_id", "")
            normalized["saved_at"] = current_timestamp_display()
            normalized["submitted_at"] = None
            bucket["records"].append(normalized)
            record_map[serial_no] = normalized
            imported += 1

    for institute, bucket in records_by_institute.items():
        save_records(bucket["records"], institute)
    return jsonify({"status": "ok", "imported": imported, "updated": updated, "total": imported + updated})

@app.route("/api/delete/<serial_no>", methods=["DELETE"])
@admin_required
def delete_record(serial_no):
    records_source = list_supabase_records() if is_supabase_enabled() else load_records()
    record_to_delete = next((r for r in records_source if r.get("serial_no") == serial_no), None)
    if not record_to_delete:
        return jsonify({"error": "Record not found"}), 404
    institute = canonicalize_institute_name(record_to_delete.get("institute_name"))
    if is_supabase_enabled():
        supabase_request(
            "DELETE",
            "records",
            query={"serial_no": f"eq.{serial_no}", "institute_name": f"eq.{institute}"},
        )
    else:
        records = load_records(institute)
        records = [r for r in records if r.get("serial_no") != serial_no]
        save_records(records, institute)
    if record_to_delete and record_to_delete.get("photo_drive_id"):
        delete_drive_file(record_to_delete.get("photo_drive_id"))
    delete_uploaded_file_from_url((record_to_delete or {}).get("photo_url"))
    delete_local_file_if_exists(os.path.join(UPLOAD_DIR, f"{serial_no}.jpg"))
    return jsonify({"status": "deleted"})

@app.route("/api/export-excel")
@admin_required
def export_excel():
    records, institute_filter = get_filtered_records()
    buf = io.BytesIO(build_export_xls_bytes(records))
    buf.seek(0)
    institute = institute_filter or (records[0].get("institute_name", "IDCardRecords") if records else "IDCardRecords")
    institute_safe = "".join(c for c in institute if c.isalnum() or c in "_ -")
    return send_file(buf, as_attachment=True,
                     download_name=f"{institute_safe}_IDCards_{datetime.now().strftime('%Y%m%d')}.xls",
                     mimetype="application/vnd.ms-excel")

@app.route("/api/export-zip")
@admin_required
def export_zip():
    records, institute_filter = get_filtered_records()
    buf = io.BytesIO()
    csv_text = build_export_csv_text(records)
    excel_bytes = build_export_xls_bytes(records)
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for index, rec in enumerate(records, 1):
            serial = rec.get("serial_no", "")
            photo_filename = f"{index}.jpg"
            photo_written = False
            photo_path = os.path.join(UPLOAD_DIR, f"{serial}.jpg")
            if os.path.exists(photo_path):
                zf.write(photo_path, f"photos/{photo_filename}")
                photo_written = True
            elif rec.get("photo_drive_id"):
                photo_bytes = download_drive_file(rec.get("photo_drive_id"))
                if photo_bytes:
                    zf.writestr(f"photos/{photo_filename}", photo_bytes)
                    photo_written = True
            elif rec.get("photo_url"):
                photo_bytes = download_file_from_url(rec.get("photo_url"))
                if photo_bytes:
                    zf.writestr(f"photos/{photo_filename}", photo_bytes)
                    photo_written = True
            if not photo_written:
                zf.writestr(f"photos/{photo_filename}", build_placeholder_avatar_bytes(rec, index))
        zf.writestr("records.csv", csv_text)
        zf.writestr("records.xls", excel_bytes)
    buf.seek(0)
    institute = institute_filter or (records[0].get("institute_name", "IDCardRecords") if records else "IDCardRecords")
    institute_safe = "".join(c for c in institute if c.isalnum() or c in "_ -")
    return send_file(buf, as_attachment=True,
                     download_name=f"{institute_safe}_IDCards_Export_{datetime.now().strftime('%Y%m%d')}.zip",
                     mimetype="application/zip")

if __name__ == "__main__":
    app.run(debug=True, port=5050)
